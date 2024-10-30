import numpy as np
import customtkinter as ctk
from tkinter import filedialog, simpledialog, messagebox
import os
import pystray
from PIL import Image
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
from openpyxl.formatting.rule import CellIsRule
import warnings

warnings.filterwarnings("ignore")
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.table import Table
import shutil
import logging

plt.rcParams['figure.max_open_warning'] = 500
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")


class MyApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.file_path1 = None
        self.folder_path = None
        self.title("Plot Draw App")
        self.geometry("700x700")
        self.protocol("WM_DELETE_WINDOW", self.minimize_to_tray)

        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure(2, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.grid_rowconfigure(2, weight=1)
        self.grid_rowconfigure(3, weight=1)

        self.sidebar_frame = ctk.CTkFrame(self, width=50, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=17, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(4, weight=1)
        self.logo_label = ctk.CTkLabel(self.sidebar_frame, text="Setting Pane  30/10/2024", anchor="center",
                                       font=ctk.CTkFont(size=15))
        self.logo_label.grid(row=3, column=0, padx=20, pady=(5, 10))

        self.file_label1 = ctk.CTkLabel(self, text="", font=('calibri', 12),
                                        justify="center", wraplength=200)
        self.file_label1.grid(row=0, column=1, columnspan=2, padx=10, pady=5, sticky="nsew")
        self.select_button1 = ctk.CTkButton(self, text="Select Excel File (.xlsx)", font=('calibri', 12),
                                            command=self.select_excel_file)
        self.select_button1.grid(row=1, column=1, columnspan=2, padx=10, pady=5, sticky="nsew", ipadx=5)

        self.folder_label = ctk.CTkLabel(self, text="", font=('calibri', 12), justify="center",
                                         wraplength=200)
        self.folder_label.grid(row=2, column=1, columnspan=2, padx=10, pady=(5, 5), sticky="nsew")
        self.select_folder_button = ctk.CTkButton(self, text="Select Output Folder", font=('calibri', 12),
                                                  command=self.select_folder)
        self.select_folder_button.grid(row=3, column=1, columnspan=2, padx=10, pady=(5, 5), sticky="nsew", ipadx=3)
        self.file_label1 = ctk.CTkLabel(self, text="", font=('calibri', 12),
                                        justify="center", wraplength=200)
        self.file_label1.grid(row=4, column=1, columnspan=2, padx=10, pady=5, sticky="nsew")

        self.integer_label = ctk.CTkLabel(self, text="Number of Last Indexes", font=('calibri', 12),
                                          justify="center",
                                          wraplength=200)
        self.integer_label.grid(row=7, column=1, padx=10, pady=(5, 5), sticky="w")
        self.number_last_index = ctk.CTkEntry(self, font=('calibri', 12), width=100)
        self.number_last_index.grid(row=7, column=2, padx=10, pady=(5, 5), sticky="w", ipadx=20)

        self.check_04_06 = ctk.BooleanVar()
        self.check_04_06_checkbox = ctk.CTkCheckBox(self, text="Check [0.4 - 0.6]", font=('calibri', 12),
                                                    variable=self.check_04_06)
        self.check_04_06_checkbox.grid(row=8, column=1, padx=10, pady=(5, 5), sticky="w")

        self.integer_label = ctk.CTkLabel(self, text="Max Number of Points", font=('calibri', 12),
                                          justify="center",
                                          wraplength=200)
        self.integer_label.grid(row=9, column=1, padx=10, pady=(5, 5), sticky="w")
        self.max_points = ctk.CTkEntry(self, font=('calibri', 12), width=100)
        self.max_points.grid(row=9, column=2, padx=10, pady=(5, 5), sticky="w", ipadx=20)

        self.custom_limits_var = ctk.BooleanVar()
        self.custom_limits_checkbox = ctk.CTkCheckBox(self, text="Custom Limits", font=('calibri', 12),
                                                      variable=self.custom_limits_var,
                                                      command=self.toggle_limit_entries)
        self.custom_limits_checkbox.grid(row=10, column=1, padx=10, pady=(5, 5), sticky="w")

        self.integer_label = ctk.CTkLabel(self, text="Enter Min Limit", font=('calibri', 12),
                                          justify="center",
                                          wraplength=200)
        self.integer_label.grid(row=11, column=1, padx=10, pady=(5, 5), sticky="w")
        self.min_limit = ctk.CTkEntry(self, font=('calibri', 12), width=100, state="disabled")
        self.min_limit.grid(row=11, column=2, padx=10, pady=(5, 5), sticky="w", ipadx=20)

        self.integer_label = ctk.CTkLabel(self, text="Enter Max Limit", font=('calibri', 12),
                                          justify="center",
                                          wraplength=200)
        self.integer_label.grid(row=12, column=1, padx=10, pady=(5, 5), sticky="w")
        self.max_limit = ctk.CTkEntry(self, font=('calibri', 12), width=100, state="disabled")
        self.max_limit.grid(row=12, column=2, padx=10, pady=(5, 5), sticky="w", ipadx=20)

        self.name_label = ctk.CTkLabel(self, text="Enter Folder Name", font=('calibri', 12),
                                       justify="center", wraplength=200)

        self.name_label.grid(row=13, column=1, padx=10, pady=(5, 5), sticky="w")
        self.folder_name = ctk.CTkEntry(self, font=('calibri', 12), width=100)
        self.folder_name.grid(row=13, column=2, padx=10, pady=(5, 5), sticky="w", ipadx=20)

        self.integer_label = ctk.CTkLabel(self, text="Enter Rows Number", font=('calibri', 12),
                                          justify="center",
                                          wraplength=200)
        self.integer_label.grid(row=14, column=1, padx=10, pady=(5, 5), sticky="w")
        self.rows_number = ctk.CTkEntry(self, font=('calibri', 12), width=100)
        self.rows_number.grid(row=14, column=2, padx=10, pady=(5, 5), sticky="w", ipadx=20)

        self.name_label = ctk.CTkLabel(self, text="Enter the Date", font=('calibri', 12),
                                       justify="center", wraplength=200)

        self.name_label.grid(row=15, column=1, padx=10, pady=(5, 5), sticky="w")
        self.filter_date = ctk.CTkEntry(self, font=('calibri', 12), width=100)
        self.filter_date.grid(row=15, column=2, padx=10, pady=(5, 5), sticky="w", ipadx=20)

        self.start_button = ctk.CTkButton(self, text="Start", font=('Tahoma', 10, 'bold'), command=self.start)
        self.start_button.grid(row=16, column=1, columnspan=2, padx=10, pady=50, sticky="nsew")

        self.appearance_mode_label = ctk.CTkLabel(self.sidebar_frame, text="Appearance Mode:", anchor="w")
        self.appearance_mode_label.grid(row=13, column=0, padx=20, pady=(10, 0))
        self.appearance_mode_optionemenu = ctk.CTkOptionMenu(self.sidebar_frame, values=["dark", "light"],
                                                             command=self.change_appearance_mode_event)
        self.appearance_mode_optionemenu.grid(row=14, column=0, padx=20, pady=(10, 10))

        self.scaling_label = ctk.CTkLabel(self.sidebar_frame, text="UI Scaling:", anchor="w")
        self.scaling_label.grid(row=15, column=0, padx=20, pady=(10, 0))
        self.scaling_optionemenu = ctk.CTkOptionMenu(self.sidebar_frame, values=["80%", "90%", "100%", "110%", "120%"],
                                                     command=self.change_scaling_event)
        self.scaling_optionemenu.grid(row=16, column=0, padx=20, pady=(10, 20))

        self.icon = None

    def ask_password(self):
        PASSWORD = "Hadi1234"

        input_password = simpledialog.askstring("Password", "Enter Password:", show='*')
        if input_password == PASSWORD:
            return True
        else:
            messagebox.showerror("Error", "Incorrect password. The application will close.")
            return False

    def select_excel_file(self):
        self.file_path1 = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])

    def select_folder(self):
        self.folder_path = filedialog.askdirectory()

    def change_appearance_mode_event(self, new_appearance_mode: str):
        ctk.set_appearance_mode(new_appearance_mode)

    def change_scaling_event(self, new_scaling: str):
        new_scaling_float = int(new_scaling.replace("%", "")) / 100
        ctk.set_widget_scaling(new_scaling_float)

    def minimize_to_tray(self):
        self.withdraw()
        image = Image.open("icon.ico")
        menu = pystray.Menu(
            pystray.MenuItem('Quit', self.quit_window),
            pystray.MenuItem('Show', self.show_window)
        )
        self.icon = pystray.Icon("name", image, "File Content Copier", menu)
        self.icon.run()

    def quit_window(self, icon, item):
        icon.stop()
        self.destroy()

    def show_window(self, icon, item):
        icon.stop()
        self.after(0, self.deiconify)

    def toggle_limit_entries(self):
        if self.custom_limits_var.get():
            self.min_limit.configure(state="normal")
            self.max_limit.configure(state="normal")
        else:
            self.min_limit.configure(state="disabled")
            self.max_limit.configure(state="disabled")

    # Function to generate sequences of indexes
    def print_sequences(self, all_rows, new_rows):
        all_combinations = []
        for start in range(1, all_rows):
            for end in range(start + 1, all_rows + 1):
                combination = list(range(start, end))
                if len(combination) < 3:
                    continue
                all_combinations.append(combination)

        new_combinations = []
        for start in range(1, all_rows - new_rows):
            for end in range(start + 1, all_rows - new_rows + 1):
                combination = list(range(start, end))
                if len(combination) < 3:
                    continue
                new_combinations.append(combination)

        final_combinations = []
        for combination in all_combinations:
            if combination not in new_combinations:
                final_combinations.append(combination)

        return final_combinations

    # Function to check if all regression line values lie within the range [0.4, 0.6]
    def check_regression_in_range(self, df, low=0, high=1):
        x = df['index']
        b_linear = np.poly1d(np.polyfit(x, df['b'], 1))(x)
        b_poly = np.poly1d(np.polyfit(x, df['b'], 2))(x)
        s_linear = np.poly1d(np.polyfit(x, df['s'], 1))(x)
        s_poly = np.poly1d(np.polyfit(x, df['s'], 2))(x)

        return all((low <= b_linear) & (b_linear <= high)) and \
            all((low <= b_poly) & (b_poly <= high)) and \
            all((low <= s_linear) & (s_linear <= high)) and \
            all((low <= s_poly) & (s_poly <= high))

    @staticmethod
    def custom_operations(group):
        # Sort the group by 'Points'
        sorted_group = group.sort_values(by='Points')

        # Initialize counters for smaller and larger areas
        smaller_count = 0
        larger_count = 0

        # Compare each row with the next one
        for i in range(len(sorted_group) - 1):
            if sorted_group['Special'].iloc[i + 1] < sorted_group['Special'].iloc[i]:
                smaller_count += 1
            elif sorted_group['Special'].iloc[i + 1] > sorted_group['Special'].iloc[i]:
                larger_count += 1

        # Assign counts to all rows in the group
        sorted_group['Not'] = smaller_count
        sorted_group['OK'] = larger_count

        return sorted_group

    # Function to plot the dataframe
    def plot_df(self, df, item, folder_name="plots", min_limit=0.4, max_limit=0.6):

        # if len(df) < 3:
        #     return None, None, None, None

        if not self.check_regression_in_range(df, min_limit, max_limit):
            return None, None, None, None

        # plt.figure(figsize=(12, 6))
        # ax = plt.gca()

        x = df['index']

        # Calculate regression lines
        b_linear = np.poly1d(np.polyfit(x, df['b'], 1))(x)
        b_poly = np.poly1d(np.polyfit(x, df['b'], 2))(x)
        s_linear = np.poly1d(np.polyfit(x, df['s'], 1))(x)
        s_poly = np.poly1d(np.polyfit(x, df['s'], 2))(x)

        if b_poly[1] > b_poly[0]:
            optimum_b = max(b_poly)
        else:
            optimum_b = min(b_poly)

        if s_poly[1] > s_poly[0]:
            optimum_s = max(s_poly)
        else:
            optimum_s = min(s_poly)

        # small_limit = round(min(optimum_b, optimum_s), 4)
        # large_limit = round(max(optimum_b, optimum_s), 4)
        final_optimum = round(abs(optimum_b - optimum_s), 4)

        # Analyze relationships between lines
        b_max_distance = max(abs(b_linear - b_poly))
        s_max_distance = max(abs(s_linear - s_poly))

        label = "Default"
        if b_max_distance <= 0.03 and s_max_distance <= 0.03:
            # Check if b lines intersect s lines
            b_min = min(min(b_linear), min(b_poly))
            b_max = max(max(b_linear), max(b_poly))
            s_min = min(min(s_linear), min(s_poly))
            s_max = max(max(s_linear), max(s_poly))

            if b_max < s_min or s_max < b_min:
                label = "Parallel"
            else:
                intersections = 0
                for i in range(1, len(x)):
                    if b_linear[i - 1] <= s_linear[i - 1] and b_linear[i] > s_linear[i]:
                        intersections += 1
                    if b_linear[i - 1] >= s_linear[i - 1] and b_linear[i] < s_linear[i]:
                        intersections += 1
                    if b_poly[i - 1] <= s_poly[i - 1] and b_poly[i] > s_poly[i]:
                        intersections += 1
                    if b_poly[i - 1] >= s_poly[i - 1] and b_poly[i] < s_poly[i]:
                        intersections += 1
                    if b_linear[i - 1] <= s_poly[i - 1] and b_linear[i] > s_poly[i]:
                        intersections += 1
                    if s_linear[i - 1] <= b_poly[i - 1] and s_linear[i] > b_poly[i]:
                        intersections += 1

                if intersections <= 3:
                    label = "One Point"
                elif intersections > 3:
                    label = "Two Points"
        else:
            # Check if b lines intersect s lines
            b_min = min(min(b_linear), min(b_poly))
            b_max = max(max(b_linear), max(b_poly))
            s_min = min(min(s_linear), min(s_poly))
            s_max = max(max(s_linear), max(s_poly))

            if b_max < s_min or s_max < b_min:
                label = "Default"
            else:
                intersections = 0
                for i in range(1, len(x)):
                    if b_linear[i - 1] <= s_linear[i - 1] and b_linear[i] > s_linear[i]:
                        intersections += 1
                    if b_linear[i - 1] >= s_linear[i - 1] and b_linear[i] < s_linear[i]:
                        intersections += 1
                    if b_poly[i - 1] <= s_poly[i - 1] and b_poly[i] > s_poly[i]:
                        intersections += 1
                    if b_poly[i - 1] >= s_poly[i - 1] and b_poly[i] < s_poly[i]:
                        intersections += 1
                    if b_linear[i - 1] <= s_poly[i - 1] and b_linear[i] > s_poly[i]:
                        intersections += 1
                    if s_linear[i - 1] <= b_poly[i - 1] and s_linear[i] > b_poly[i]:
                        intersections += 1

                if intersections > 3:
                    label = "Two Points"

        if label:

            # sns.regplot(x='index', y='b', data=df, scatter=False,
            #             line_kws={'color': '#FFD700', 'linestyle': (0, (1, 1))},
            #             ci=None, label='Linear Regression b')
            # sns.regplot(x='index', y='b', data=df, scatter=False,
            #             line_kws={'color': '#FF1493', 'linestyle': (0, (1, 1))},
            #             order=2, ci=None, label='Polynomial Regression b')
            # sns.regplot(x='index', y='s', data=df, scatter=False, line_kws={'color': 'black', 'linestyle': (0, (1, 1))},
            #             ci=None, label='Linear Regression s')
            # sns.regplot(x='index', y='s', data=df, scatter=False,
            #             line_kws={'color': '#00FFFF', 'linestyle': (0, (1, 1))},
            #             order=2, ci=None, label='Polynomial Regression s')
            #
            # plt.title('Regression Plot')
            # plt.xlabel('Index')
            # plt.ylabel('Value')
            # plt.grid(True)
            # plt.ylim(0, 1)
            # plt.legend()

            # Add table to the plot
            # Points = len(df)
            # Date = df.iloc[-1]['Date']
            # Time = df.iloc[-1]['Time']
            # Density_b = round(df.iloc[-1]['b'], 6)
            # Density_s = round(df.iloc[-1]['s'], 6)
            # Measure = round(df.iloc[-1]['H'] - df.iloc[-1]['I'], 3)
            # Sum_Density_b = round(sum(df['b']), 6)
            # Sum_Density_s = round(sum(df['s']), 6)
            # Density_Diff = round(round(sum(df['b']), 6) - round(sum(df['s']), 6), 6)
            # Color = round(df.iloc[-1]['C'] - df.iloc[-1]['O'], 3)
            # Mean_Color = round((sum(df['C']) - sum(df['O'])) / len(df), 3)
            # CM = round(Color / Measure, 4)

            # data_table = [['Points', Points], ['Date', Date], ['Time', Time], ['Density_b', Density_b],
            #               ['Density_s', Density_s],
            #               ['Measure', Measure], ['Sum_Density_b', Sum_Density_b], ['Sum_Density_s', Sum_Density_s],
            #               ['Color', Color], ['Mean_Color', Mean_Color], ['Density_Diff', Density_Diff],
            #               ['Min_Limit', small_limit], ['Max_Limit', large_limit], ['Difference', final_optimum],
            #               ['C/M', CM]]

            # table = Table(ax, bbox=[-0.3, 0.22, 0.23, 0.55])
            # for i, (key, value) in enumerate(data_table):
            #     table.add_cell(i, 0, width=1.0, height=0.2, text=key, loc='left', edgecolor='black')
            #     table.add_cell(i, 1, width=1.0, height=0.2, text=value, loc='right', edgecolor='black')
            # table.set_fontsize(20)
            # table.scale(1, 1.5)
            # ax.add_table(table)

            main_path = os.path.join(self.folder_path, folder_name) if folder_name else self.folder_path
            if not os.path.exists(main_path):
                os.makedirs(main_path)
            # plt.tight_layout()

            # copy_df = pd.read_excel(self.file_path1, header=None)
            # copy_df = copy_df.iloc[:, :6]
            # copy_df.to_excel(f"{main_path}/data.xlsx", index=False, header=False)

            child_folder_1 = f"[{min_limit}_{max_limit}]"
            child_folder_2 = f"[{str(df.iloc[0]['index'])}_{str(df.iloc[len(df) - 1]['index'])}]_{str(df.iloc[len(df) - 1]['Time']).replace(':', ';')[0:-3] + "_" + str(df.iloc[len(df) - 1]['Date'])}"

            # if child_folder_2 in previous_plots_list:
            #     return None, None, None, None
            save_path = os.path.join(main_path, child_folder_1, label, child_folder_2)
            # all_path = os.path.join(main_path, child_folder_1, label, 'all')
            # if not os.path.exists(save_path):
            #     os.makedirs(save_path)
            #
            # if not os.path.exists(all_path):
            #     os.makedirs(all_path)

            fig_name = f"{save_path}/[{item[0]}_{item[-1]}]_{df.iloc[-1]['Date']}_{str(df.iloc[-1]['Time']).replace(':', ';')[0:-3]}_[{min_limit}_{max_limit}].png"

            x_fit = np.linspace(0, len(b_linear) - 1, len(b_linear))
            all_y_fits = np.vstack((b_linear, b_poly, s_linear, s_poly))

            # Calculate max and min across all y_fit arrays
            max_y_fit = np.max(all_y_fits, axis=0)
            min_y_fit = np.min(all_y_fits, axis=0)

            # Plot max and min lines
            # plt.figure(figsize=(16, 6))

            # plt.plot(x_fit, b_linear, '--', color='red', label='B Linear')
            # plt.plot(x_fit, b_poly, '-', color='blue', label='B Polynomial')
            # plt.plot(x_fit, s_linear, '--', color='purple', label='S Linear')
            # plt.plot(x_fit, s_poly, '-', color='orange', label='S Polynomial')

            # Fill the area between max and min lines
            # plt.fill_between(x_fit, min_y_fit, max_y_fit, color='lightgray', alpha=0.5, label='Filled Area')

            # Calculate the area using the trapezoidal rule
            area = np.trapezoid(max_y_fit - min_y_fit, x_fit)

            # Add plot labels and title
            # plt.title('Maximum Area Across Lines')
            # plt.xlabel('X values')
            # plt.ylabel('Y values')
            # # plt.ylim(0, 1)
            # plt.legend(loc='upper left', bbox_to_anchor=(1, 1))
            # plt.grid(True)
            # plt.tight_layout(rect=[0, 0, 0.75, 1])

            # Display the calculated area on the plot
            # plt.text(1.05, 0.3, f'Filled Area: {area:.5f}', transform=plt.gca().transAxes)

            # plt.savefig(fig_name)

            # all_fig_name = f"{all_path}/[{item[0]}_{item[-1]}]_{df.iloc[-1]['Date']}_{str(df.iloc[-1]['Time']).replace(':', ';')[0:-3]}_[{min_limit}_{max_limit}].png"
            # plt.savefig(all_fig_name)
            #
            # plt.close()

            # df.to_excel(
            #     f"{save_path}/[{item[0]}_{item[-1]}]_{df.iloc[-1]['Date']}_{str(df.iloc[-1]['Time']).replace(':', ';')[0:-3]}_[{min_limit}_{max_limit}].xlsx",
            #     index=False)

            return child_folder_2, final_optimum, fig_name, area
        return None, None, None, None

    def start(self):
        if not self.file_path1 or not self.folder_path:
            messagebox.showerror("Error", "Please select an Excel file and an output folder.")
            return

        # Continue with existing start method logic
        if self.number_last_index.get() != "":
            number_last_index = int(self.number_last_index.get())
        if self.min_limit.get() != "":
            min_limit = self.min_limit.get().split("_")
            min_limit_list = [float(x) for x in min_limit]
        if self.max_limit.get() != "":
            max_limit = self.max_limit.get().split("_")
            max_limit_list = [float(x) for x in max_limit]
        if self.folder_name.get() != "":
            folder_name = self.folder_name.get()

        if self.custom_limits_var.get():
            if self.min_limit.get() != "":
                min_limit = self.min_limit.get().split("_")
                min_limit_list = [float(x) for x in min_limit]
            if self.max_limit.get() != "":
                max_limit = self.max_limit.get().split("_")
                max_limit_list = [float(x) for x in max_limit]
        else:
            if self.check_04_06.get():
                min_limit_list = [0.47, 0.46, 0.45, 0.4]
                max_limit_list = [0.57, 0.58, 0.59, 0.6]
            else:
                min_limit_list = [0.47, 0.46, 0.45]
                max_limit_list = [0.57, 0.58, 0.59]

        if self.folder_name.get() == "":
            folder_name = "plots"

        df = pd.read_excel(self.file_path1, header=None)
        # added these 4 line
        now = pd.Timestamp.now().strftime("%Y-%m-%d %H-%M-%S")
        main_path = os.path.join(self.folder_path, folder_name) if folder_name else self.folder_path
        if not os.path.exists(main_path):
            os.makedirs(main_path)

        df.to_excel(f"{main_path}\\sample_{now}.xlsx", index=False)

        df = df.iloc[:, :6]
        if self.number_last_index.get() == "":
            number_last_index = len(df)

        df.columns = ['Date', 'Time', 'O', 'H', 'I', 'C']
        df['index'] = df.index
        df['index'] = range(1, len(df) + 1)
        df = df[['index', 'Date', 'Time', 'O', 'H', 'I', 'C']]

        df['b'] = (df['C'] - df['I']) / (df['H'] - df['I'])
        df['s'] = (df['H'] - df['C']) / (df['H'] - df['I'])

        if self.max_points.get() == "":
            max_points = len(df)
        else:
            max_points = int(self.max_points.get())

        for i in range(len(min_limit_list)):
            previous_plots = []
            all_figs = []

            all_dates = []
            all_times = []
            all_density_b = []
            all_density_s = []
            all_measure = []
            all_color = []
            temp_areas = []
            all_points = []

            first_min_limit = min_limit_list[i]
            first_max_limit = max_limit_list[i]
            start = len(df) + 1
            end = number_last_index
            sequences = self.print_sequences(start, end)

            for item in sequences:
                temp_df = df.loc[df['index'].isin(item)]
                if len(temp_df) > max_points:
                    continue
                child_folder, final_difference, fig_name, area = self.plot_df(temp_df, item, folder_name,
                                                                              min_limit=first_min_limit,
                                                                              max_limit=first_max_limit)
                if child_folder and final_difference:
                    previous_plots.append(child_folder)
                    temp_list = [final_difference, fig_name]
                    all_figs.append(temp_list)
                    temp_date = temp_df.iloc[-1]['Date']
                    temp_time = temp_df.iloc[-1]['Time']
                    temp_density_b = round(temp_df.iloc[-1]['b'], 6)
                    temp_density_s = round(temp_df.iloc[-1]['s'], 6)
                    temp_measure = round(temp_df.iloc[-1]['H'] - temp_df.iloc[-1]['I'], 3)
                    temp_color = round(temp_df.iloc[-1]['C'] - temp_df.iloc[-1]['O'], 3)
                    all_dates.append(temp_date)
                    all_times.append(temp_time)
                    all_density_b.append(temp_density_b)
                    all_density_s.append(temp_density_s)
                    all_measure.append(temp_measure)
                    all_color.append(temp_color)
                    temp_areas.append(area / len(temp_df))
                    all_points.append(len(temp_df))

            # if len(temp_areas) > 0:
            #     min_areas.append(min(temp_areas))
            #     mean_areas.append(sum(temp_areas))
            #     all_points.append(len(temp_areas))

            sorted_path = os.path.join(self.folder_path, folder_name, 'sorted')
            if not os.path.exists(sorted_path):
                os.makedirs(sorted_path)

            all_df = pd.DataFrame(
                {'Date': all_dates, 'Time': all_times, 'Density_b': all_density_b, 'Density_s': all_density_s,
                 'Measure': all_measure, 'Color': all_color, 'Special': temp_areas, 'Points': all_points})

            if len(all_df) == 0:
                continue
            # all_df = pd.DataFrame(
            #         {'Date': all_dates, 'Time': all_times, })

            # Sort all_df by Date
            all_df['Date'] = pd.to_datetime(all_df['Date'])
            all_df = all_df.sort_values(by='Date')
            now = pd.Timestamp.now().strftime("%Y-%m-%d %H-%M-%S")
            all_df.to_excel(f"{sorted_path}\\all_{now}_{first_min_limit}_{first_max_limit}.xlsx", index=False)

            all_df = all_df.sort_values(by='Date', ascending=False)
            all_df['Date_Time'] = all_df['Date'].astype(str) + '_' + all_df['Time'].astype(str)
            time_counts = all_df['Date_Time'].value_counts().sort_index()
            all_df['Number_of_repeat'] = all_df['Date_Time'].map(time_counts)
            df_grouped = all_df.groupby(['Time', 'Date'])['Special'].agg(Minimum_Area='min',
                                                                         Total_Area='sum').reset_index()
            df_merged = pd.merge(all_df, df_grouped, on=['Time', 'Date'], how='left')

            df_merged = df_merged.groupby(['Date', 'Time']).apply(MyApp.custom_operations).reset_index(drop=True)

            df_unique = df_merged.drop_duplicates(subset='Date_Time').sort_values(by='Number_of_repeat',
                                                                                  ascending=False).reset_index(
                drop=True)
            df_unique = df_unique[
                ['Date', 'Number_of_repeat', 'Time', 'Density_b', 'Density_s', 'Measure', 'Color', 'Minimum_Area',
                 'Total_Area', 'Not', 'OK']]
            # df_unique = df_unique[['Date', 'Number_of_repeat', 'Time', 'Minimum_Area', 'Total_Area']]
            df_sorted = df_unique.sort_values(by=['Date', 'Number_of_repeat'], ascending=[False, False])
            df_sorted['Date'] = df_sorted['Date'].astype(str)

            df_2 = df_sorted[['Date', 'Time', 'Color', 'Not', 'OK', 'Minimum_Area']]
            column_name = str(first_min_limit) + '_' + str(first_max_limit)
            df_2[column_name] = abs(df_2['Not'] - df_2['OK'])
            df_2['r.e'] = df_2.apply(lambda row: row[column_name] * (-1) if row['Color'] < 0 else row[column_name],
                                     axis=1)

            df_2 = df_2[['Date', 'Time', 'Color', column_name, 'r.e', 'Minimum_Area']]

            if len(df_sorted) > 0:
                sorted_now = pd.Timestamp.now().strftime("%Y-%m-%d %H-%M-%S").split(" ")[1] + "_" + folder_name
                final_path = f"{sorted_path}\\{str(df_sorted['Date'].iloc[0]).split(" ")[0]}_{sorted_now}_{first_min_limit}_{first_max_limit}.xlsx"

                with pd.ExcelWriter(final_path, engine='xlsxwriter') as writer:
                    # Write df_1 to sheet_1
                    df_sorted.to_excel(writer, sheet_name='Sorted', index=False)

                    # Write df_2 to sheet_2
                    df_2.to_excel(writer, sheet_name='Violation', index=False)
                if os.path.exists(final_path):
                    workbook = load_workbook(final_path)
                    sheet = workbook['Violation']

                    # Apply conditional formatting to the "Color" column (let's say it's column C)
                    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                    green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

                    # Assuming "Color" is in column C (third column), and data starts from row 2
                    col_index = 3
                    for row in range(2, sheet.max_row + 1):
                        cell = sheet.cell(row=row, column=col_index)
                        if cell.value < 0:
                            cell.fill = red_fill
                        elif cell.value >= 0:
                            cell.fill = green_fill

                    # Save the workbook
                    workbook.save(final_path)
                else:
                    print("File does not exist or wasn't saved properly.")

                # df_sorted.to_excel(f"{sorted_path}\\{str(df_sorted['Date'].iloc[0]).split(" ")[0]}_{sorted_now}.xlsx",
                #                    index=False)

            # if all_figs:
            #     all_figs.sort(key=lambda x: x[0])
            #     fig_counter = 1
            #     for item in all_figs:
            #         fig_name = item[1]
            #         fig_name = fig_name.split('/')[-1]
            #         copy_fig_path = f"{sorted_path}/{fig_counter}_{fig_name}"
            #         shutil.copy(item[1], copy_fig_path)
            #         fig_counter += 1
            #
        current_year = str(datetime.now().year)
        if self.rows_number.get() != "":
            rows_number = int(self.rows_number.get())
        else:
            rows_number = 15

        abstract_df = pd.DataFrame()
        pivot_date = []
        pivot_time = []
        pivot_total = []
        pivot_minimum_areas = []
        pivot_color = []
        for item in os.listdir(sorted_path):
            if item.endswith('.xlsx') or item.endswith('.xls'):
                if item.startswith(current_year):
                    excel_file_path = os.path.join(sorted_path, item)
                    temp_excel = pd.read_excel(excel_file_path, sheet_name='Violation')
                    if self.filter_date.get() != "":
                        temp_excel = temp_excel[temp_excel['Date'] == self.filter_date.get()]

                    minimum_row_temp = min(len(temp_excel), rows_number)
                    pivot_date.extend(temp_excel['Date'].tolist()[:minimum_row_temp])
                    pivot_time.extend(temp_excel['Time'].tolist()[:minimum_row_temp])
                    pivot_total.extend(temp_excel.iloc[:, 3].tolist()[:minimum_row_temp])
                    pivot_minimum_areas.extend(temp_excel['Minimum_Area'].tolist()[:minimum_row_temp])
                    pivot_color.extend(temp_excel['Color'].tolist()[:minimum_row_temp])

                    if len(temp_excel) >= rows_number:
                        selected_temp_df = temp_excel.head(rows_number)
                        selected_temp_df.loc[rows_number + 3] = [""] * len(selected_temp_df.columns)
                        selected_temp_df.loc[rows_number + 4] = [""] * len(selected_temp_df.columns)
                        selected_temp_df.loc[rows_number + 5] = [""] * len(selected_temp_df.columns)
                        selected_temp_df.loc[rows_number + 6] = [""] * len(selected_temp_df.columns)
                        selected_temp_df.loc[rows_number + 7] = [""] * len(selected_temp_df.columns)
                        selected_temp_df['Color'] = pd.to_numeric(selected_temp_df['Color'], errors='coerce')
                        selected_temp_df['r.e'] = pd.to_numeric(selected_temp_df['r.e'], errors='coerce')
                        sum_r_e = (selected_temp_df['r.e']).sum()

                        positive_color = (selected_temp_df['Color'] >= 0).sum()
                        negative_color = (selected_temp_df['Color'] < 0).sum()
                        all_colors_values = positive_color + negative_color
                        selected_temp_df.at[rows_number + 3, 'r.e'] = sum_r_e
                        selected_temp_df.at[rows_number + 5, 'Time'] = 'Positive'
                        selected_temp_df.at[rows_number + 5, 'r.e'] = round((positive_color / all_colors_values), 3)
                        selected_temp_df.at[rows_number + 6, 'Time'] = 'Negative'
                        selected_temp_df.at[rows_number + 6, 'r.e'] = round((negative_color / all_colors_values), 3)
                        selected_temp_df.at[rows_number + 7, 'Time'] = 'Sum'
                        selected_temp_df.at[rows_number + 5, 'Color'] = positive_color
                        selected_temp_df.at[rows_number + 6, 'Color'] = negative_color
                        selected_temp_df.at[rows_number + 7, 'Color'] = positive_color + negative_color

                        empty_column = pd.DataFrame([''] * rows_number, columns=['DEF'])
                        abstract_df = pd.concat([abstract_df, selected_temp_df, empty_column], axis=1)
                    else:
                        empty_column = pd.DataFrame([''] * rows_number, columns=['DEF'])
                        selected_temp_df = temp_excel
                        for row_index in range(len(temp_excel), rows_number+1):
                            selected_temp_df.loc[row_index] = [""] * len(selected_temp_df.columns)

                        selected_temp_df.loc[rows_number + 3] = [""] * len(selected_temp_df.columns)
                        selected_temp_df.loc[rows_number + 4] = [""] * len(selected_temp_df.columns)
                        selected_temp_df.loc[rows_number + 5] = [""] * len(selected_temp_df.columns)
                        selected_temp_df.loc[rows_number + 6] = [""] * len(selected_temp_df.columns)
                        selected_temp_df.loc[rows_number + 7] = [""] * len(selected_temp_df.columns)
                        selected_temp_df['Color'] = pd.to_numeric(selected_temp_df['Color'], errors='coerce')
                        selected_temp_df['r.e'] = pd.to_numeric(selected_temp_df['r.e'], errors='coerce')
                        sum_r_e = (selected_temp_df['r.e']).sum()
                        positive_color = (selected_temp_df['Color'] >= 0).sum()
                        negative_color = (selected_temp_df['Color'] < 0).sum()
                        all_colors_values = positive_color + negative_color
                        selected_temp_df.at[rows_number + 3, 'r.e'] = sum_r_e
                        selected_temp_df.at[rows_number + 5, 'Time'] = 'Positive'
                        selected_temp_df.at[rows_number + 5, 'r.e'] = round((positive_color / all_colors_values), 3)
                        selected_temp_df.at[rows_number + 6, 'Time'] = 'Negative'
                        selected_temp_df.at[rows_number + 6, 'r.e'] = round((negative_color / all_colors_values), 3)
                        selected_temp_df.at[rows_number + 7, 'Time'] = 'Sum'
                        selected_temp_df.at[rows_number + 5, 'Color'] = positive_color
                        selected_temp_df.at[rows_number + 6, 'Color'] = negative_color
                        selected_temp_df.at[rows_number + 7, 'Color'] = positive_color + negative_color

                        abstract_df = pd.concat([abstract_df, selected_temp_df, empty_column], axis=1)

        # number_of_limits = len(abstract_df.columns) // 7
        # for i in range(number_of_limits - 1):
        #     next_column_index = (7 * (i + 1)) + 4
        #     before_column_index = (7 * i) + 4
        #     def_percentage_index = (7 * i) + 6
        #     abstract_df.iloc[rows_number + 4, def_percentage_index] = 'DEF'
        #     abstract_df.iloc[rows_number + 5, def_percentage_index] = abstract_df.iloc[rows_number + 2, next_column_index] - \
        #                                                         abstract_df.iloc[rows_number + 2, before_column_index]
        #     abstract_df.iloc[rows_number + 6, def_percentage_index] = abstract_df.iloc[rows_number + 3, next_column_index] - \
        #                                                         abstract_df.iloc[rows_number + 3, before_column_index]

        # abstract_df = abstract_df.iloc[:, :-1]
        # make a dataframe for the pivot table
        pivot_df = pd.DataFrame({
            'date': pivot_date,
            'time': pivot_time,
            'total': pivot_total,
            'minimum_area': pivot_minimum_areas,
            'color': pivot_color
        })
        pivot_df['date_time'] = pivot_df['date'] + "_" + pivot_df['time']
        grouped_pivot_df = pivot_df.groupby('date_time').agg(
            {'total': 'sum', 'time': lambda x: ', '.join(set(x)), 'date_time': 'count',
             'date': lambda x: ', '.join(set(x)), 'minimum_area': 'mean',
             'color': 'mean'}).rename(columns={'date_time': 'num rep'})
        # sort by total
        grouped_pivot_df = grouped_pivot_df.sort_values(by='total', ascending=False)

        grouped_pivot_df = grouped_pivot_df.reset_index()

        grouped_pivot_df = grouped_pivot_df[['date', 'time', 'total', 'num rep', 'minimum_area', 'color']]

        violaion_file_name = f"{sorted_path}\\violation_{now}.xlsx"

        with pd.ExcelWriter(violaion_file_name, engine='openpyxl') as writer:
            abstract_df.to_excel(writer, sheet_name='abstract', index=False)
            grouped_pivot_df.to_excel(writer, sheet_name='pivot', index=False)

        # abstract_df.to_excel(f"{sorted_path}\\violation_{now}.xlsx", sheet_name='abstract', index=False)

        colored_last_row = min(len(temp_excel), rows_number) + 1

        # Load the workbook and select the active worksheet
        wb = load_workbook(violaion_file_name)
        ws_abstract = wb['abstract']
        ws_pivot = wb['pivot']

        # Define the color fills for negative (red) and positive (green) values
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

        # Get the header row to find the 'color' columns
        header_abstract = [cell.value for cell in ws_abstract[1]]

        # Apply conditional formatting to all 'color' columns
        for col_idx, column_name in enumerate(header_abstract, start=1):
            if 'color' in column_name.lower():  # Check if the column name contains 'color'
                # Apply red fill for negative values
                # {'endsWith', 'beginsWith', 'notContains', 'lessThanOrEqual', 'containsText', 'lessThan', 'greaterThanOrEqual', 'notBetween', 'equal', 'greaterThan', 'between', 'notEqual'}
                ws_abstract.conditional_formatting.add(
                    f"{ws_abstract.cell(2, col_idx).coordinate}:{ws_abstract.cell(colored_last_row, col_idx).coordinate}",
                    CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))
                # Apply green fill for positive values
                ws_abstract.conditional_formatting.add(
                    f"{ws_abstract.cell(2, col_idx).coordinate}:{ws_abstract.cell(colored_last_row, col_idx).coordinate}",
                    CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill))
            if 'r.e' in column_name.lower():
                col_values = [ws_abstract.cell(row=row, column=col_idx).value for row in
                              range(2, ws_abstract.max_row + 1)]
                col_sum = sum([val for val in col_values if isinstance(val, (int, float))])
                if col_sum < 0:  # Apply conditional formatting if sum is negative
                    # Apply red fill to all cells in this column (rows 2 to the last row)
                    ws_abstract.conditional_formatting.add(
                        f"{ws_abstract.cell(2, col_idx).coordinate}:{ws_abstract.cell(colored_last_row, col_idx).coordinate}",
                        CellIsRule(operator='lessThanOrEqual', formula=['10000'], fill=red_fill)
                    )
                else:
                    # Apply green fill to all cells in this column (rows 2 to the last row)
                    ws_abstract.conditional_formatting.add(
                        f"{ws_abstract.cell(2, col_idx).coordinate}:{ws_abstract.cell(colored_last_row, col_idx).coordinate}",
                        CellIsRule(operator='greaterThanOrEqual', formula=['-10000'], fill=green_fill))

        for row in range(2, ws_pivot.max_row + 1):
            color_value = ws_pivot.cell(row=row, column=6).value
            time_cell = ws_pivot.cell(row=row, column=2)
            if color_value <= 0:
                time_cell.fill = PatternFill(start_color="FF9999", end_color="FF9999",
                                             fill_type="solid")
            else:
                time_cell.fill = PatternFill(start_color="99FF99", end_color="99FF99",
                                             fill_type="solid")

        for row in range(2, ws_pivot.max_row):
            total_value_one = ws_pivot.cell(row=row, column=3).value
            total_value_two = ws_pivot.cell(row=row + 1, column=3).value
            if total_value_one == total_value_two:
                ws_pivot.cell(row=row, column=3).fill = PatternFill(start_color="C7A7DB", end_color="C7A7DB",
                                                                    fill_type="solid")
                ws_pivot.cell(row=row + 1, column=3).fill = PatternFill(start_color="C7A7DB", end_color="C7A7DB",
                                                                        fill_type="solid")

        # Save the updated workbook with conditional formatting

        color_col_index = None
        for cell in wb['pivot'][1]:
            if cell.value == 'color':
                color_col_index = cell.column

        if color_col_index:
            wb['pivot'].delete_cols(color_col_index, 1)

        wb.save(violaion_file_name)

        messagebox.showinfo("Process Completed", "The process has finished successfully!")


if __name__ == "__main__":
    # ask for password
    if not MyApp().ask_password():
        exit()
    app = MyApp()
    app.mainloop()
