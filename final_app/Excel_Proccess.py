import matplotlib.pyplot as plt
from tkinter import Tk
from tkinter.filedialog import askopenfilename, asksaveasfilename
import tkinter as tk
from tkinter import ttk, messagebox
import threading
import os
import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import traceback
import re

logging.basicConfig(filename='app.log',level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s',encoding='utf-8')

def select_file(title="لطفاً فایل اکسل را انتخاب کنید", filetypes=[("Excel files", "*.xlsx;*.xls")]):
    """نمایش پنجره برای انتخاب فایل اکسل"""
    Tk().withdraw()
    return askopenfilename(title=title, filetypes=filetypes)
    """نمایش پنجره برای انتخاب فایل اکسل و ثبت خطاها در صورت وجود"""
    try:
        Tk().withdraw()  # مخفی کردن پنجره اصلی Tkinter
        file_path = askopenfilename(title=title, filetypes=filetypes)

        if not file_path:
            logging.info("کاربر فایلی انتخاب نکرده است.")
            checkbox.config(state="normal")
            return None

        logging.info(f"فایل انتخاب شده: {file_path}")
        return file_path
    except Exception as e:
        logging.error(f"خطا در انتخاب فایل: {e}")
        return None

def save_file(title="فایل ذخیره‌سازی را انتخاب کنید", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")]):
    """نمایش پنجره برای ذخیره فایل اکسل"""
    Tk().withdraw()
    return asksaveasfilename(defaultextension=defaultextension, filetypes=filetypes, title=title)

def process_excel_file(file_path):
    """خواندن و پردازش داده‌های فایل اکسل"""
    try:
        # خواندن فایل اکسل
        df = pd.read_excel(file_path, header=None)
        if df.empty:
            logging.error(f"فایل معرفی شده حاوی هیچ اطلاعاتی نمی باشد")
        else:
            pattern = r'\b\d{4}\.\d{2}\.\d{2}\b'
            matches = re.findall(pattern, df[0].iloc[0])
            if len(matches) == 0:
                df = df.drop(0).reset_index(drop=True)
            # حذف سرستون‌ها و قرار دادن سرستون جدید
            new_header = ['date', 'time', 'open', 'high', 'low', 'close']
            df.columns = new_header
            df['date'] = df['date'].astype(str)
            df['time'] = df['time'].astype(str)
            return df

    except Exception as e:
        logging.error(f"خطای غیرمنتظره در پردازش فایل: {e}")
        logging.error(traceback.format_exc())

def calculate_area_ratio(X1, Y1, X2, Y2, X3, Y3, x_perp, y_perp):
    try:
        # محاسبه مساحت اول
        area_1 = abs(X1 * (Y2 - y_perp) + X2 * (y_perp - Y1) + x_perp * (Y1 - Y2)) / 2
        # محاسبه مساحت دوم
        area_2 = abs(X3 * (Y2 - y_perp) + X2 * (y_perp - Y3) + x_perp * (Y3 - Y2)) / 2
        # محاسبه نسبت مساحت‌ها
        if area_2 != 0:  # جلوگیری از تقسیم بر صفر
            ratio = area_1 / area_2
        else:
            ratio = 0  # اگر مساحت دوم صفر باشد، نسبت مشخص نمی‌شود
        return area_1, area_2, ratio
    except Exception as e:
        logging.error(f"خطای غیرمنتظره در محاسبه مساحت‌ها: {e}")

def calculate_perpendicular_distance(X1, Y1, X2, Y2, X3, Y3):
    try:
        # محاسبه شیب خط
        if X3 == X1:  # جلوگیری از تقسیم بر صفر در محاسبه شیب
            raise ValueError("شیب خط نامشخص است (X1 و X3 برابر هستند).")

        m = (Y3 - Y1) / (X3 - X1)
        A = m
        B = -1
        C = Y1 - m * X1
        value = X1 * (Y3 - Y2) + X3 * (Y2 - Y1) + X2 * (Y1 - Y3)
        if value > 0:
            YPC = 1
        elif value < 0:
            YPC = -1
        else:
            YPC = 0  # نقطه روی خط
        # پارامتر فاصله عمودی
        d = (A * X2 + B * Y2 + C) / (A ** 2 + B ** 2)
        x_perp = X2 - A * d
        y_perp = Y2 - B * d
        return x_perp, y_perp, m, YPC

    except ZeroDivisionError:
        logging.error("خطای تقسیم بر صفر: X1 و X3 برابر هستند.")
    except ValueError as ve:
        logging.error(f"خطای مقدار نادرست: {ve}")
    except Exception as e:
        logging.error(f"خطای غیرمنتظره در محاسبه فاصله عمودی: {e}")

def plot_triangle_with_perpendicular(X1, Y1, l1, X2, Y2, l2, X3, Y3, l3, x_perp, y_perp, date):
    try :
        if checkbox_var.get():
            # تعیین مقادیر محور x و y برای تنظیم محدوده نمودار
            folder_path = 'pics'
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)
                logging.info(f"پوشه '{folder_path}' ایجاد شد.")
            else:
                pass

            x_values = [X1, X2, X3, X1]
            y_values = [Y1, Y2, Y3, Y1]

            x_min = min(x_values)
            x_max = max(x_values)
            y_min = min(y_values)
            y_max = max(y_values)

            # رسم نمودار
            plt.figure(figsize=(8, 8))
            plt.plot([X3, x_perp], [Y3, y_perp], 'g--', label='perpendicular line of Point 2')
            plt.plot(x_perp, y_perp, 'mo', label='intersection point')

            plt.plot(x_values, y_values, marker='o', linestyle='-', color='b')
            plt.text(X1, Y1, 'Point ' + l1, fontsize=10, verticalalignment='bottom', horizontalalignment='right')
            plt.text(X2, Y2, 'Point ' + l2, fontsize=10, verticalalignment='bottom', horizontalalignment='right')
            plt.text(X3, Y3, 'Point ' + l3, fontsize=10, verticalalignment='bottom', horizontalalignment='right')

            plt.xlim(x_min * 0.999, x_max * 1.001)
            plt.ylim(y_min * 0.999, y_max * 1.001)
            plt.xlabel('times')
            plt.ylabel('prices')
            plt.title('Triangle Plot')
            plt.legend()
            plt.grid(True)
            # ذخیره تصویر
            plt.savefig('pics/' + date + '.png', dpi=300, bbox_inches='tight')
            plt.close()
    except Exception as e:
        logging.error(f"خطای غیرمنتظره در رسم نمودار مثلث: {e}")

def analyze_data(df, progress_var, progress_callback):
    try :
        """تحلیل داده‌ها و ایجاد نمودارهای مربوطه"""
        list_of_result = []
        unique_dates = df['date'].drop_duplicates().tolist()
        logging.info(f"تعداد تاریخ‌ها : {len(unique_dates)}")

        for counter, item in enumerate(unique_dates, start=1):
            date = item
            try:
                progress_callback(counter / len(unique_dates) * 100)
                full_data_of_date = df[df['date'] == item].copy()
                full_data_of_date.reset_index(drop=True, inplace=True)

                if len(full_data_of_date) >= 14 and len(full_data_of_date) < 23 :

                    first_time = full_data_of_date['time'].iloc[0]
                    if first_time == "00:30:00":
                        full_data_of_date.at[0, 'time'] = '01:30:00'

                    if (full_data_of_date['time'].str.contains('01:30:00').any() and
                            full_data_of_date['time'].str.contains('02:30:00').any() and
                            full_data_of_date['time'].str.contains('03:30:00').any() and
                            full_data_of_date['time'].str.contains('04:30:00').any() and
                            full_data_of_date['time'].str.contains('05:30:00').any() and
                            full_data_of_date['time'].str.contains('06:30:00').any() and
                            full_data_of_date['time'].str.contains('07:30:00').any() and
                            full_data_of_date['time'].str.contains('08:30:00').any() and
                            full_data_of_date['time'].str.contains('09:30:00').any() and
                            full_data_of_date['time'].str.contains('10:30:00').any() and
                            full_data_of_date['time'].str.contains('11:30:00').any() and
                            full_data_of_date['time'].str.contains('12:30:00').any() and
                            full_data_of_date['time'].str.contains('13:30:00').any() and
                            full_data_of_date['time'].str.contains('14:30:00').any()):

                        data_0130_1430 = full_data_of_date.iloc[:14].copy()
                        data_0130_1430['close-open'] = data_0130_1430['close'] - data_0130_1430['open']

                        max_close_0130_1430 = data_0130_1430['close'].max()
                        max_close_index_0130_1430 = data_0130_1430['close'].idxmax()
                        min_close_0130_1430 = data_0130_1430['close'].min()
                        min_close_index_0130_1430 = data_0130_1430['close'].idxmin()

                        Y1 = data_0130_1430[data_0130_1430['time'] == "01:30:00"]["close"].iloc[0]
                        X1 = data_0130_1430[data_0130_1430['time'] == "01:30:00"]["close"].index[0] + 1
                        Y3 = data_0130_1430[data_0130_1430['time'] == "14:30:00"]["close"].iloc[0]
                        X3 = data_0130_1430[data_0130_1430['time'] == "14:30:00"]["close"].index[0] + 1

                        if round(abs(Y1 - min_close_0130_1430), 2) > round(abs(Y1 - max_close_0130_1430), 2):
                            Y2 = min_close_0130_1430
                            X2 = min_close_index_0130_1430 + 1
                        elif round(abs(Y1 - min_close_0130_1430), 2) < round(abs(Y1 - max_close_0130_1430), 2):
                            Y2 = max_close_0130_1430
                            X2 = max_close_index_0130_1430 + 1

                        if Y2 == Y3 and X2 == X3:
                            open_1430 = data_0130_1430[data_0130_1430['time'] == "14:30:00"]["open"].iloc[0]
                            if open_1430 > Y2:
                                Y2 = open_1430
                            else:
                                Y2 = Y3
                                Y3 = open_1430

                        X_list = [X1, X2, X3]
                        Y_list = [Y1, Y2, Y3]
                        x_max = max(X_list)
                        x_min = min(X_list)
                        y_max = max(Y_list)
                        y_min = min(Y_list)
                        i = 1
                        for x in X_list:
                            if i == 1:
                                X1 = y_min + (x - x_min) * (y_max - y_min) / (x_max - x_min)
                            elif i == 2:
                                X2 = y_min + (x - x_min) * (y_max - y_min) / (x_max - x_min)
                            elif i == 3:
                                X3 = y_min + (x - x_min) * (y_max - y_min) / (x_max - x_min)
                            i += 1

                        flag = False
                        x_perp, y_perp, m, YPC = calculate_perpendicular_distance(X1, Y1, X2, Y2, X3, Y3)
                        if x_perp <= X3 and x_perp >= X1:
                            area_1, area_2, ratio = calculate_area_ratio(X1, Y1, X2, Y2, X3, Y3, x_perp, y_perp)
                            PDP = 2
                            plot_triangle_with_perpendicular(X1, Y1, "1", X3, Y3, "3", X2, Y2, "2", x_perp, y_perp, date)
                            flag = True

                        if flag == False:
                            x_perp, y_perp, m, YPC = calculate_perpendicular_distance(X1, Y1, X3, Y3, X2, Y2)
                            if x_perp <= X2 and x_perp >= X1:
                                area_1, area_2, ratio = calculate_area_ratio(X1, Y1, X3, Y3, X2, Y2, x_perp, y_perp)
                                PDP = 3
                                plot_triangle_with_perpendicular(X1, Y1, "1", X2, Y2, "2", X3, Y3, "3", x_perp, y_perp,date)
                            else:
                                x_perp, y_perp, m, YPC = calculate_perpendicular_distance(X2, Y2, X1, Y1, X3, Y3)
                                area_1, area_2, ratio = calculate_area_ratio(X2, Y2, X1, Y1, X3, Y3, x_perp, y_perp)
                                PDP = 1
                                plot_triangle_with_perpendicular(X3, Y3, "3", X2, Y2, "2", X1, Y1, "1", x_perp, y_perp,date)

                        CO_1530 = 0
                        CO_1630 = 0
                        CO_1730 = 0
                        CO_1830 = 0
                        CO_1930 = 0
                        CO_2030 = 0
                        CO_2130 = 0
                        CO_2230 = 0
                        CO_2330 = 0

                        # افزودن نتایج به لیست
                        list_of_result.append([date,
                                               round(ratio, 2),
                                               PDP,
                                               round(m, 2),
                                               0,
                                               0,
                                               Y3,
                                               CO_1530,
                                               CO_1630,
                                               CO_1730,
                                               CO_1830,
                                               CO_1930,
                                               CO_2030,
                                               CO_2130,
                                               CO_2230,
                                               CO_2330,
                                               Y1,
                                               Y2,
                                               Y3,
                                               round(area_1, 2),
                                               round(area_2, 2),
                                               0,
                                               0,
                                               YPC])
                    else :
                        logging.info(f"داده {counter} در تاریخ {item} از {len(unique_dates)} کمتر از طول نرمال است.")

                elif len(full_data_of_date) < 23:
                    logging.info(f"داده {counter} در تاریخ {item} از {len(unique_dates)} کمتر از طول نرمال است.")
                    continue
                elif len(full_data_of_date) == 23 :

                    full_data_of_date['close-open'] = full_data_of_date['close'] - full_data_of_date['open']
                    # ادامه کد برای تحلیل و رسم نمودار
                    data_0130_1430 = full_data_of_date.iloc[:14].copy()
                    data_1530_2330 = full_data_of_date.iloc[14:23].copy()

                    first_time = data_0130_1430['time'].iloc[0]
                    if first_time == "00:30:00":
                        data_0130_1430.at[0, 'time'] = '01:30:00'

                    max_close_0130_1430 = data_0130_1430['close'].max()
                    max_close_index_0130_1430 = data_0130_1430['close'].idxmax()
                    min_close_0130_1430 = data_0130_1430['close'].min()
                    min_close_index_0130_1430 = data_0130_1430['close'].idxmin()

                    Y1 = data_0130_1430[data_0130_1430['time'] == "01:30:00"]["close"].iloc[0]
                    X1 = data_0130_1430[data_0130_1430['time'] == "01:30:00"]["close"].index[0] + 1
                    Y3 = data_0130_1430[data_0130_1430['time'] == "14:30:00"]["close"].iloc[0]
                    X3 = data_0130_1430[data_0130_1430['time'] == "14:30:00"]["close"].index[0] + 1

                    if round(abs(Y1 - min_close_0130_1430), 2) > round(abs(Y1 - max_close_0130_1430), 2):
                        Y2 = min_close_0130_1430
                        X2 = min_close_index_0130_1430 + 1
                    elif round(abs(Y1 - min_close_0130_1430), 2) < round(abs(Y1 - max_close_0130_1430), 2):
                        Y2 = max_close_0130_1430
                        X2 = max_close_index_0130_1430 + 1

                    if Y2 == Y3 and X2 == X3:
                        open_1430 = data_0130_1430[data_0130_1430['time'] == "14:30:00"]["open"].iloc[0]
                        if open_1430 > Y2:
                            Y2 = open_1430
                        else:
                            Y2 = Y3
                            Y3 = open_1430

                    max_close_1530_2330 = data_1530_2330['close'].max()
                    min_close_1530_2330 = data_1530_2330['close'].min()
                    CLOSE_2330 = data_1530_2330[data_1530_2330['time'] == "23:30:00"]["close"].iloc[0]

                    X_list = [X1, X2, X3]
                    Y_list = [Y1, Y2, Y3]
                    x_max = max(X_list)
                    x_min = min(X_list)
                    y_max = max(Y_list)
                    y_min = min(Y_list)
                    i = 1
                    for x in X_list:
                        if i == 1:
                            X1 = y_min + (x - x_min) * (y_max - y_min) / (x_max - x_min)
                        elif i == 2:
                            X2 = y_min + (x - x_min) * (y_max - y_min) / (x_max - x_min)
                        elif i == 3:
                            X3 = y_min + (x - x_min) * (y_max - y_min) / (x_max - x_min)
                        i += 1

                    flag = False
                    x_perp, y_perp, m, YPC = calculate_perpendicular_distance(X1, Y1, X2, Y2, X3, Y3)
                    if x_perp <= X3 and x_perp >= X1:
                        area_1, area_2, ratio = calculate_area_ratio(X1, Y1, X2, Y2, X3, Y3, x_perp, y_perp)
                        PDP = 2
                        plot_triangle_with_perpendicular(X1, Y1,"1", X3, Y3,"3", X2, Y2,"2", x_perp, y_perp, date)
                        flag = True

                    if flag == False:
                        x_perp, y_perp, m, YPC = calculate_perpendicular_distance(X1, Y1, X3, Y3, X2, Y2)
                        if x_perp <= X2 and x_perp >= X1:
                            area_1, area_2, ratio = calculate_area_ratio(X1, Y1, X3, Y3, X2, Y2, x_perp, y_perp)
                            PDP = 3
                            plot_triangle_with_perpendicular(X1, Y1,"1", X2, Y2,"2", X3, Y3,"3", x_perp, y_perp, date)
                        else:
                            x_perp, y_perp, m, YPC = calculate_perpendicular_distance(X2, Y2, X1, Y1, X3, Y3)
                            area_1, area_2, ratio = calculate_area_ratio(X2, Y2, X1, Y1, X3, Y3, x_perp, y_perp)
                            PDP = 1
                            plot_triangle_with_perpendicular(X3, Y3,"3", X2, Y2,"2", X1, Y1,"1", x_perp, y_perp, date)

                    CO_1530 = round(data_1530_2330[data_1530_2330['time'] == "15:30:00"]["close-open"].iloc[0], 2)
                    CO_1630 = round(data_1530_2330[data_1530_2330['time'] == "16:30:00"]["close-open"].iloc[0], 2)
                    CO_1730 = round(data_1530_2330[data_1530_2330['time'] == "17:30:00"]["close-open"].iloc[0], 2)
                    CO_1830 = round(data_1530_2330[data_1530_2330['time'] == "18:30:00"]["close-open"].iloc[0], 2)
                    CO_1930 = round(data_1530_2330[data_1530_2330['time'] == "19:30:00"]["close-open"].iloc[0], 2)
                    CO_2030 = round(data_1530_2330[data_1530_2330['time'] == "20:30:00"]["close-open"].iloc[0], 2)
                    CO_2130 = round(data_1530_2330[data_1530_2330['time'] == "21:30:00"]["close-open"].iloc[0], 2)
                    CO_2230 = round(data_1530_2330[data_1530_2330['time'] == "22:30:00"]["close-open"].iloc[0], 2)
                    CO_2330 = round(data_1530_2330[data_1530_2330['time'] == "23:30:00"]["close-open"].iloc[0], 2)

                    # افزودن نتایج به لیست
                    list_of_result.append([date,
                                           round(ratio, 2),
                                           PDP,
                                           round(m, 2),
                                           round(CLOSE_2330 - Y3, 2),
                                           round(CLOSE_2330 - Y1, 2),
                                           Y3,
                                           CO_1530,
                                           CO_1630,
                                           CO_1730,
                                           CO_1830,
                                           CO_1930,
                                           CO_2030,
                                           CO_2130,
                                           CO_2230,
                                           CO_2330,
                                           Y1,
                                           Y2,
                                           Y3,
                                           round(area_1, 2),
                                           round(area_2, 2),
                                           round(min_close_1530_2330, 2),
                                           round(max_close_1530_2330, 2),
                                           YPC])
            except Exception as e:
                logging.error(f"خطای غیرمنتظره در آنالیز دیتا {date, e}")
                logging.error(traceback.format_exc())
                continue

        return list_of_result
    except Exception as e:
        logging.error(f"خطای غیرمنتظره در آنالیز{date, e}")

def save_results_to_excel(df, file_path):
    try :
        """ذخیره نتایج پردازش‌شده در فایل اکسل"""
        if not file_path:
            logging.info("فایلی برای ذخیره‌سازی انتخاب نشده است.")
            return
        # معکوس کردن ترتیب ردیف‌ها
        df_reversed = df.iloc[::-1].reset_index(drop=True)
        df_reversed.to_excel(file_path, index=False, engine='openpyxl')
        workbook = load_workbook(file_path)
        # افزودن قالب‌بندی شرطی اگر نیاز است
        sheet = workbook.active
        # تعیین رنگ‌ها
        red_fill = PatternFill(start_color="FF6666", end_color="FF0000", fill_type="solid")
        green_fill = PatternFill(start_color="66FF66", end_color="00FF00", fill_type="solid")

        for row in range(2, sheet.max_row + 1):  # فرض می‌کنیم ردیف اول عنوان است
            cell = sheet[f"{'X'}{row}"]
            cell1 = sheet[f"{'B'}{row}"]
            cell3 = sheet[f"{'V'}{row}"]
            cell4 = sheet[f"{'W'}{row}"]
            cell3.fill = red_fill
            cell4.fill = green_fill
            if cell.value is not None:  # بررسی اینکه سلول خالی نباشد
                if cell.value < 0:
                    cell.fill = red_fill
                    cell1.fill = red_fill
                elif cell.value > 0:
                    cell.fill = green_fill
                    cell1.fill = green_fill

        column_letters = ['D', 'E', 'F', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']

        for col in column_letters:
            # پیمایش سلول‌ها در ستون و اعمال رنگ‌ها
            for row in range(2, sheet.max_row + 1):  # فرض می‌کنیم ردیف اول عنوان است
                cell = sheet[f'{col}{row}']
                if cell.value is not None:  # بررسی اینکه سلول خالی نباشد
                    if cell.value < 0:
                        cell.fill = red_fill
                    elif cell.value > 0:
                        cell.fill = green_fill
        workbook.save(file_path)
        logging.info(f"فایل با موفقیت در مسیر {file_path} ذخیره شد.")
    except Exception as e:
        logging.error(f"خطای غیرمنتظره در ذخیره سازی فایل excel {e}")

def update_progress(value):
    progress_var.set(value)
    root.update_idletasks()

def main():
    file_path = select_file()
    checkbox.config(state="disabled")
    df = process_excel_file(file_path)
    if df is not None:
        results = analyze_data(df, progress_var, update_progress)
        if results:
            column_names = ['date',
                            'ratio',
                            'PDP',
                            'chord slope',
                            'mid result',
                            'f result',
                            'C 14:30',
                            'C-O 15:30',
                            'C-O 16:30',
                            'C-O 17:30',
                            'C-O 18:30',
                            'C-O 19:30',
                            'C-O 20:30',
                            'C-O 21:30',
                            'C-O 22:30',
                            'C-O 23:30',
                            'point1',
                            'point2',
                            'point3',
                            'area1',
                            'area2',
                            'min',
                            'max',
                            'YPC']
            result_df = pd.DataFrame(results, columns=column_names)
            result_df.fillna(0, inplace=True)
            save_path = save_file()
            save_results_to_excel(result_df, save_path)
            checkbox.config(state="normal")
            messagebox.showinfo("پیغام", "پردازش با موفقیت به اتمام رسید")
        else:
            checkbox.config(state="normal")
    else:
        checkbox.config(state="normal")

try :
    root = tk.Tk()
    root.title("Progress Bar")
    window_width = 400
    window_height = 150
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    position_top = int(screen_height/2 - window_height/2)
    position_right = int(screen_width/2 - window_width/2)
    root.geometry(f"{window_width}x{window_height}+{position_right}+{position_top}")
    progress_var = tk.DoubleVar()
    progress_bar = ttk.Progressbar(root, variable=progress_var, maximum=100, length=350)
    progress_bar.pack(pady=20)
    start_button = ttk.Button(root, text="انتخاب فایل", command=lambda: threading.Thread(target=main, daemon=True).start())
    start_button.pack()
    checkbox_var = tk.BooleanVar()
    checkbox = ttk.Checkbutton(root, text="اگر نیاز به ذخیره فایل تصویری دارید", variable=checkbox_var)
    checkbox.pack(pady=10)
    root.mainloop()
except Exception as e:
    logging.error(e)