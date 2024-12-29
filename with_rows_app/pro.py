import numpy as np
import customtkinter as ctk
from tkinter import filedialog, simpledialog, messagebox
import os
import pystray
from PIL import Image
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from collections import Counter
import warnings
import logging
import inspect

warnings.filterwarnings("ignore")
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")


class MyApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.file_path1 = None
        self.folder_path = None
        self.title("with rows1")
        self.geometry("540x470")
        self.protocol("WM_DELETE_WINDOW", self.minimize_to_tray)

        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure(2, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.grid_rowconfigure(2, weight=1)
        self.grid_rowconfigure(3, weight=1)

        self.sidebar_frame = ctk.CTkFrame(self, width=50, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=11, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(4, weight=1)
        self.logo_label = ctk.CTkLabel(self.sidebar_frame, text="Setting Pane  27/12/2024", anchor="center",
                                       font=ctk.CTkFont(size=15))
        self.logo_label.grid(row=2, column=0, padx=20, pady=(5, 10))

        self.select_button = ctk.CTkButton(self, text="Select Excel File (.xlsx)", font=('calibri', 12),
                                           command=self.select_excel_file)
        self.select_button.grid(row=1, column=1, columnspan=2, padx=10, pady=5, sticky="nsew", ipadx=5)

        self.select_folder_button = ctk.CTkButton(self, text="Select Output Folder", font=('calibri', 12),
                                                  command=self.select_folder)
        self.select_folder_button.grid(row=2, column=1, columnspan=2, padx=10, pady=(5, 5), sticky="nsew", ipadx=3)

        self.integer_label = ctk.CTkLabel(self, text="Enter Min Limit", font=('calibri', 12),
                                          justify="center",
                                          wraplength=200)
        self.integer_label.grid(row=3, column=1, padx=10, pady=(5, 5), sticky="w")
        self.min_limit = ctk.CTkEntry(self, font=('calibri', 12), width=100, state="normal")
        self.min_limit.grid(row=3, column=2, padx=10, pady=(5, 5), sticky="w", ipadx=20)

        self.integer_label = ctk.CTkLabel(self, text="Enter Max Limit", font=('calibri', 12),
                                          justify="center",
                                          wraplength=200)
        self.integer_label.grid(row=4, column=1, padx=10, pady=(5, 5), sticky="w")
        self.max_limit = ctk.CTkEntry(self, font=('calibri', 12), width=100, state="normal")
        self.max_limit.grid(row=4, column=2, padx=10, pady=(5, 5), sticky="w", ipadx=20)

        self.name_label = ctk.CTkLabel(self, text="Enter Folder Name", font=('calibri', 12),
                                       justify="center", wraplength=200)

        self.name_label.grid(row=5, column=1, padx=10, pady=(5, 5), sticky="w")
        self.folder_name = ctk.CTkEntry(self, font=('calibri', 12), width=100)
        self.folder_name.grid(row=5, column=2, padx=10, pady=(5, 5), sticky="w", ipadx=20)

        self.integer_label = ctk.CTkLabel(self, text="Starting Point", font=('calibri', 12),
                                          justify="center",
                                          wraplength=200)
        self.integer_label.grid(row=7, column=1, padx=10, pady=(5, 5), sticky="w")

        self.rows_number = ctk.CTkEntry(self, font=('calibri', 12), width=100)
        self.rows_number.grid(row=7, column=2, padx=10, pady=(5, 5), sticky="w", ipadx=20)

        self.label = ctk.CTkLabel(self, text="Start Row", font=("calibri", 12), justify="center",
                                  wraplength=200)
        self.label.grid(row=8, column=1, padx=10, pady=(5, 5), sticky="w")
        self.start_row = ctk.CTkEntry(self, font=('calibri', 12), width=100)
        self.start_row.grid(row=8, column=2, padx=10, pady=(5, 5), sticky="w", ipadx=20)

        self.label = ctk.CTkLabel(self, text="End Row", font=("calibri", 12), justify="center",
                                  wraplength=200)
        self.label.grid(row=9, column=1, padx=10, pady=(5, 5), sticky="w")

        self.data_rows = ctk.CTkEntry(self, font=('calibri', 12), width=100)
        self.data_rows.grid(row=9, column=2, padx=10, pady=(5, 5), sticky="w", ipadx=20)

        self.start_button = ctk.CTkButton(self, text="Start", font=('Tahoma', 10, 'bold'), command=self.start)
        self.start_button.grid(row=10, column=1, columnspan=2, padx=10, pady=50, sticky="nsew")

        self.appearance_mode_label = ctk.CTkLabel(self.sidebar_frame, text="Appearance Mode:", anchor="w")
        self.appearance_mode_label.grid(row=6, column=0, padx=20, pady=(10, 0))
        self.appearance_mode_optionemenu = ctk.CTkOptionMenu(self.sidebar_frame, values=["dark", "light"],
                                                             command=self.change_appearance_mode_event)
        self.appearance_mode_optionemenu.grid(row=7, column=0, padx=20, pady=(10, 10))

        self.scaling_label = ctk.CTkLabel(self.sidebar_frame, text="UI Scaling:", anchor="w")
        self.scaling_label.grid(row=8, column=0, padx=20, pady=(10, 0))
        self.scaling_optionemenu = ctk.CTkOptionMenu(self.sidebar_frame, values=["80%", "90%", "100%", "110%", "120%"],
                                                     command=self.change_scaling_event)
        self.scaling_optionemenu.grid(row=9, column=0, padx=20, pady=(10, 20))

        self.icon = None

    def ask_password(self):
        PASSWORD = "Hadi1234"
        input_password = simpledialog.askstring("Password", "Enter Password:", show='*')
        if input_password == PASSWORD:
            return True
        else:
            messagebox.showerror("Error", "Incorrect password. The application will close.")
            return False

    def select_excel_file(self):
        self.file_path1 = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])

    def calculate_area_ratio(self, X1, Y1, X2, Y2, X3, Y3, x_perp, y_perp):
        try:
            area_1 = abs(X1 * (Y2 - y_perp) + X2 * (y_perp - Y1) + x_perp * (Y1 - Y2)) / 2
            area_2 = abs(X3 * (Y2 - y_perp) + X2 * (y_perp - Y3) + x_perp * (Y3 - Y2)) / 2
            if area_2 != 0:
                ratio = area_1 / area_2
            else:
                ratio = 0
            return area_1, area_2, ratio
        except Exception as e:
            logging.error(f"There is an error in calculate area and ratio: {e}")

    def calculate_perpendicular_distance(self, X1, Y1, X2, Y2, X3, Y3):
        try:
            if X3 == X1:
                raise ValueError("شیب خط نامشخص است (X1 و X3 برابر هستند).")

            m = (Y3 - Y1) / (X3 - X1)
            A = m
            B = -1
            C = Y1 - m * X1
            value = X1 * (Y3 - Y2) + X3 * (Y2 - Y1) + X2 * (Y1 - Y3)
            if value > 0:
                YPC = 1
            elif value < 0:
                YPC = -1
            else:
                YPC = 0
            d = (A * X2 + B * Y2 + C) / (A ** 2 + B ** 2)
            x_perp = X2 - A * d
            y_perp = Y2 - B * d
            return x_perp, y_perp, m, YPC

        except ZeroDivisionError:
            logging.error("error of diving to zero X1 and X3")
        except ValueError as ve:
            logging.error(f"Error of incorrect value: {ve}")
        except Exception as e:
            logging.error(f"Unexpected error in calculate distance: {e}")

    def select_folder(self):
        self.folder_path = filedialog.askdirectory()

    def change_appearance_mode_event(self, new_appearance_mode: str):
        ctk.set_appearance_mode(new_appearance_mode)

    def change_scaling_event(self, new_scaling: str):
        new_scaling_float = int(new_scaling.replace("%", "")) / 100
        ctk.set_widget_scaling(new_scaling_float)

    def minimize_to_tray(self):
        self.withdraw()
        image = Image.open("icon.ico")
        menu = pystray.Menu(
            pystray.MenuItem('Quit', self.quit_window),
            pystray.MenuItem('Show', self.show_window)
        )
        self.icon = pystray.Icon("name", image, "File Content Copier", menu)
        self.icon.run()

    class LineNumberFormatter(logging.Formatter):
        def format(self, record):
            # Get the line number of the calling frame
            record.lineno = getattr(record, 'lineno', 'unknown')
            return super().format(record)

    def quit_window(self, icon, item):
        icon.stop()
        self.destroy()

    def show_window(self, icon, item):
        icon.stop()
        self.after(0, self.deiconify)

    # def analyze_data(self, df):
    #     try:
    #         list_of_result = []
    #         unique_dates = df['date'].drop_duplicates().tolist()
    #
    #         for counter, item in enumerate(unique_dates, start=1):
    #             date = item
    #             try:
    #                 full_data_of_date = df[df['date'] == item].copy()
    #                 full_data_of_date.reset_index(drop=True, inplace=True)
    #
    #                 if len(full_data_of_date) >= 14 and len(full_data_of_date) < 23:
    #
    #                     first_time = full_data_of_date['time'].iloc[0]
    #                     if first_time == "00:30:00":
    #                         full_data_of_date.at[0, 'time'] = '01:30:00'
    #
    #                     if (full_data_of_date['time'].str.contains('01:30:00').any() and
    #                             full_data_of_date['time'].str.contains('02:30:00').any() and
    #                             full_data_of_date['time'].str.contains('03:30:00').any() and
    #                             full_data_of_date['time'].str.contains('04:30:00').any() and
    #                             full_data_of_date['time'].str.contains('05:30:00').any() and
    #                             full_data_of_date['time'].str.contains('06:30:00').any() and
    #                             full_data_of_date['time'].str.contains('07:30:00').any() and
    #                             full_data_of_date['time'].str.contains('08:30:00').any() and
    #                             full_data_of_date['time'].str.contains('09:30:00').any() and
    #                             full_data_of_date['time'].str.contains('10:30:00').any() and
    #                             full_data_of_date['time'].str.contains('11:30:00').any() and
    #                             full_data_of_date['time'].str.contains('12:30:00').any() and
    #                             full_data_of_date['time'].str.contains('13:30:00').any() and
    #                             full_data_of_date['time'].str.contains('14:30:00').any()):
    #
    #                         data_0130_1430 = full_data_of_date.iloc[:14].copy()
    #                         data_0130_1430['close-open'] = data_0130_1430['close'] - data_0130_1430['open']
    #
    #                         max_close_0130_1430 = data_0130_1430['close'].max()
    #                         max_close_index_0130_1430 = data_0130_1430['close'].idxmax()
    #                         min_close_0130_1430 = data_0130_1430['close'].min()
    #                         min_close_index_0130_1430 = data_0130_1430['close'].idxmin()
    #
    #                         Y1 = data_0130_1430[data_0130_1430['time'] == "01:30:00"]["close"].iloc[0]
    #                         X1 = data_0130_1430[data_0130_1430['time'] == "01:30:00"]["close"].index[0] + 1
    #                         Y3 = data_0130_1430[data_0130_1430['time'] == "14:30:00"]["close"].iloc[0]
    #                         X3 = data_0130_1430[data_0130_1430['time'] == "14:30:00"]["close"].index[0] + 1
    #
    #                         if round(abs(Y1 - min_close_0130_1430), 2) > round(abs(Y1 - max_close_0130_1430), 2):
    #                             Y2 = min_close_0130_1430
    #                             X2 = min_close_index_0130_1430 + 1
    #                         elif round(abs(Y1 - min_close_0130_1430), 2) < round(abs(Y1 - max_close_0130_1430), 2):
    #                             Y2 = max_close_0130_1430
    #                             X2 = max_close_index_0130_1430 + 1
    #
    #                         if Y2 == Y3 and X2 == X3:
    #                             open_1430 = data_0130_1430[data_0130_1430['time'] == "14:30:00"]["open"].iloc[0]
    #                             if open_1430 > Y2:
    #                                 Y2 = open_1430
    #                             else:
    #                                 Y2 = Y3
    #                                 Y3 = open_1430
    #
    #                         X_list = [X1, X2, X3]
    #                         Y_list = [Y1, Y2, Y3]
    #                         x_max = max(X_list)
    #                         x_min = min(X_list)
    #                         y_max = max(Y_list)
    #                         y_min = min(Y_list)
    #                         i = 1
    #                         for x in X_list:
    #                             if i == 1:
    #                                 X1 = y_min + (x - x_min) * (y_max - y_min) / (x_max - x_min)
    #                             elif i == 2:
    #                                 X2 = y_min + (x - x_min) * (y_max - y_min) / (x_max - x_min)
    #                             elif i == 3:
    #                                 X3 = y_min + (x - x_min) * (y_max - y_min) / (x_max - x_min)
    #                             i += 1
    #
    #                         flag = False
    #                         x_perp, y_perp, m, YPC = self.calculate_perpendicular_distance(X1, Y1, X2, Y2, X3, Y3)
    #                         if x_perp <= X3 and x_perp >= X1:
    #                             area_1, area_2, ratio = self.calculate_area_ratio(X1, Y1, X2, Y2, X3, Y3, x_perp, y_perp)
    #                             PDP = 2
    #
    #                             flag = True
    #
    #                         if flag == False:
    #                             x_perp, y_perp, m, YPC = self.calculate_perpendicular_distance(X1, Y1, X3, Y3, X2, Y2)
    #                             if x_perp <= X2 and x_perp >= X1:
    #                                 area_1, area_2, ratio = self.calculate_area_ratio(X1, Y1, X3, Y3, X2, Y2, x_perp, y_perp)
    #                                 PDP = 3
    #
    #                             else:
    #                                 x_perp, y_perp, m, YPC = self.calculate_perpendicular_distance(X2, Y2, X1, Y1, X3, Y3)
    #                                 area_1, area_2, ratio = self.calculate_area_ratio(X2, Y2, X1, Y1, X3, Y3, x_perp, y_perp)
    #                                 PDP = 1
    #
    #
    #                         CO_1530 = 0
    #                         CO_1630 = 0
    #                         CO_1730 = 0
    #                         CO_1830 = 0
    #                         CO_1930 = 0
    #                         CO_2030 = 0
    #                         CO_2130 = 0
    #                         CO_2230 = 0
    #                         CO_2330 = 0
    #
    #                         # افزودن نتایج به لیست
    #                         list_of_result.append([date,
    #                                                round(ratio, 2),
    #                                                PDP,
    #                                                round(m, 2),
    #                                                0,
    #                                                0,
    #                                                Y3,
    #                                                CO_1530,
    #                                                CO_1630,
    #                                                CO_1730,
    #                                                CO_1830,
    #                                                CO_1930,
    #                                                CO_2030,
    #                                                CO_2130,
    #                                                CO_2230,
    #                                                CO_2330,
    #                                                Y1,
    #                                                Y2,
    #                                                Y3,
    #                                                round(area_1, 2),
    #                                                round(area_2, 2),
    #                                                0,
    #                                                0,
    #                                                YPC])
    #                     else:
    #                         logging.info(f"Data {counter} length is less than Normal length {len(unique_dates)} at date {item} between 14 and 23 condition.")
    #
    #                 elif len(full_data_of_date) < 23:
    #                     logging.info(f"Data {counter} length is less than Normal length {len(unique_dates)} at date {item} less than 23 condition.")
    #                     continue
    #                 elif len(full_data_of_date) == 23:
    #                     full_data_of_date['close-open'] = full_data_of_date['close'] - full_data_of_date['open']
    #                     # ادامه کد برای تحلیل و رسم نمودار
    #                     data_0130_1430 = full_data_of_date.iloc[:14].copy()
    #                     data_1530_2330 = full_data_of_date.iloc[14:23].copy()
    #
    #                     first_time = data_0130_1430['time'].iloc[0]
    #                     if first_time == "00:30:00":
    #                         data_0130_1430.at[0, 'time'] = '01:30:00'
    #
    #                     max_close_0130_1430 = data_0130_1430['close'].max()
    #                     max_close_index_0130_1430 = data_0130_1430['close'].idxmax()
    #                     min_close_0130_1430 = data_0130_1430['close'].min()
    #                     min_close_index_0130_1430 = data_0130_1430['close'].idxmin()
    #
    #                     Y1 = data_0130_1430[data_0130_1430['time'] == "01:30:00"]["close"].iloc[0]
    #                     X1 = data_0130_1430[data_0130_1430['time'] == "01:30:00"]["close"].index[0] + 1
    #                     Y3 = data_0130_1430[data_0130_1430['time'] == "14:30:00"]["close"].iloc[0]
    #                     X3 = data_0130_1430[data_0130_1430['time'] == "14:30:00"]["close"].index[0] + 1
    #
    #                     if round(abs(Y1 - min_close_0130_1430), 2) > round(abs(Y1 - max_close_0130_1430), 2):
    #                         Y2 = min_close_0130_1430
    #                         X2 = min_close_index_0130_1430 + 1
    #                     elif round(abs(Y1 - min_close_0130_1430), 2) < round(abs(Y1 - max_close_0130_1430), 2):
    #                         Y2 = max_close_0130_1430
    #                         X2 = max_close_index_0130_1430 + 1
    #
    #                     if Y2 == Y3 and X2 == X3:
    #                         open_1430 = data_0130_1430[data_0130_1430['time'] == "14:30:00"]["open"].iloc[0]
    #                         if open_1430 > Y2:
    #                             Y2 = open_1430
    #                         else:
    #                             Y2 = Y3
    #                             Y3 = open_1430
    #
    #                     max_close_1530_2330 = data_1530_2330['close'].max()
    #                     min_close_1530_2330 = data_1530_2330['close'].min()
    #                     CLOSE_2330 = data_1530_2330[data_1530_2330['time'] == "23:30:00"]["close"].iloc[0]
    #
    #                     X_list = [X1, X2, X3]
    #                     Y_list = [Y1, Y2, Y3]
    #                     x_max = max(X_list)
    #                     x_min = min(X_list)
    #                     y_max = max(Y_list)
    #                     y_min = min(Y_list)
    #                     i = 1
    #                     for x in X_list:
    #                         if i == 1:
    #                             X1 = y_min + (x - x_min) * (y_max - y_min) / (x_max - x_min)
    #                         elif i == 2:
    #                             X2 = y_min + (x - x_min) * (y_max - y_min) / (x_max - x_min)
    #                         elif i == 3:
    #                             X3 = y_min + (x - x_min) * (y_max - y_min) / (x_max - x_min)
    #                         i += 1
    #
    #                     flag = False
    #                     x_perp, y_perp, m, YPC = self.calculate_perpendicular_distance(X1, Y1, X2, Y2, X3, Y3)
    #                     if x_perp <= X3 and x_perp >= X1:
    #                         area_1, area_2, ratio = self.calculate_area_ratio(X1, Y1, X2, Y2, X3, Y3, x_perp, y_perp)
    #                         PDP = 2
    #                         flag = True
    #
    #                     if flag == False:
    #                         x_perp, y_perp, m, YPC = self.calculate_perpendicular_distance(X1, Y1, X3, Y3, X2, Y2)
    #                         if x_perp <= X2 and x_perp >= X1:
    #                             area_1, area_2, ratio = self.calculate_area_ratio(X1, Y1, X3, Y3, X2, Y2, x_perp, y_perp)
    #                             PDP = 3
    #                         else:
    #                             x_perp, y_perp, m, YPC = self.calculate_perpendicular_distance(X2, Y2, X1, Y1, X3, Y3)
    #                             area_1, area_2, ratio = self.calculate_area_ratio(X2, Y2, X1, Y1, X3, Y3, x_perp, y_perp)
    #                             PDP = 1
    #
    #                     CO_1530 = round(data_1530_2330[data_1530_2330['time'] == "15:30:00"]["close-open"].iloc[0], 2)
    #                     CO_1630 = round(data_1530_2330[data_1530_2330['time'] == "16:30:00"]["close-open"].iloc[0], 2)
    #                     CO_1730 = round(data_1530_2330[data_1530_2330['time'] == "17:30:00"]["close-open"].iloc[0], 2)
    #                     CO_1830 = round(data_1530_2330[data_1530_2330['time'] == "18:30:00"]["close-open"].iloc[0], 2)
    #                     CO_1930 = round(data_1530_2330[data_1530_2330['time'] == "19:30:00"]["close-open"].iloc[0], 2)
    #                     CO_2030 = round(data_1530_2330[data_1530_2330['time'] == "20:30:00"]["close-open"].iloc[0], 2)
    #                     CO_2130 = round(data_1530_2330[data_1530_2330['time'] == "21:30:00"]["close-open"].iloc[0], 2)
    #                     CO_2230 = round(data_1530_2330[data_1530_2330['time'] == "22:30:00"]["close-open"].iloc[0], 2)
    #                     CO_2330 = round(data_1530_2330[data_1530_2330['time'] == "23:30:00"]["close-open"].iloc[0], 2)
    #
    #
    #                     list_of_result.append([date,
    #                                            round(ratio, 2),
    #                                            PDP,
    #                                            round(m, 2),
    #                                            round(CLOSE_2330 - Y3, 2),
    #                                            round(CLOSE_2330 - Y1, 2),
    #                                            Y3,
    #                                            CO_1530,
    #                                            CO_1630,
    #                                            CO_1730,
    #                                            CO_1830,
    #                                            CO_1930,
    #                                            CO_2030,
    #                                            CO_2130,
    #                                            CO_2230,
    #                                            CO_2330,
    #                                            Y1,
    #                                            Y2,
    #                                            Y3,
    #                                            round(area_1, 2),
    #                                            round(area_2, 2),
    #                                            round(min_close_1530_2330, 2),
    #                                            round(max_close_1530_2330, 2),
    #                                            YPC])
    #             except Exception as e:
    #                 logging.error(f"An unexepted error occured: {date, e}")
    #                 continue
    #
    #         return list_of_result
    #     except Exception as e:
    #         logging.error(f"There is an error in analyzing {date, e}")

    # Function to generate sequences of indexes
    def print_sequences(self, all_rows, new_rows):
        all_combinations = []
        for start in range(1, all_rows):
            for end in range(start + 1, all_rows + 1):
                combination = list(range(start, end))
                if len(combination) < 3:
                    continue
                all_combinations.append(combination)

        new_combinations = []
        for start in range(1, all_rows - new_rows):
            for end in range(start + 1, all_rows - new_rows + 1):
                combination = list(range(start, end))
                if len(combination) < 3:
                    continue
                new_combinations.append(combination)

        final_combinations = []
        for combination in all_combinations:
            if combination not in new_combinations:
                final_combinations.append(combination)

        return final_combinations

    @staticmethod
    def custom_operations(group):
        # Sort the group by 'Points'
        sorted_group = group.sort_values(by='Points')

        # Initialize counters for smaller and larger areas
        smaller_count = 0
        larger_count = 0

        # Compare each row with the next one
        for i in range(len(sorted_group) - 1):
            if sorted_group['Special'].iloc[i + 1] < sorted_group['Special'].iloc[i]:
                smaller_count += 1
            elif sorted_group['Special'].iloc[i + 1] > sorted_group['Special'].iloc[i]:
                larger_count += 1

        # Assign counts to all rows in the group
        sorted_group['Not'] = smaller_count
        sorted_group['OK'] = larger_count

        return sorted_group

    def check_regression_in_range(self, df, low=0, high=1):
        x = df['index']
        b_linear = np.poly1d(np.polyfit(x, df['b'], 1))(x)
        b_poly = np.poly1d(np.polyfit(x, df['b'], 2))(x)
        s_linear = np.poly1d(np.polyfit(x, df['s'], 1))(x)
        s_poly = np.poly1d(np.polyfit(x, df['s'], 2))(x)

        return all((low <= b_linear) & (b_linear <= high)) and \
            all((low <= b_poly) & (b_poly <= high)) and \
            all((low <= s_linear) & (s_linear <= high)) and \
            all((low <= s_poly) & (s_poly <= high))

    def check_numbers(self, all_numbers):
        if all(num >= 0 for num in all_numbers):
            return "Green"
        elif all(num < 0 for num in all_numbers):
            return "Red"
        else:
            return "Purple"

    def update_ratio(self, row):
        if row['YPC'] < 0:
            return f"Red_{row['ratio']}"
        elif row['YPC'] > 0:
            return f"Green_{row['ratio']}"
        else:  # YPC == 0
            return f"{row['ratio']}"

    def find_min_abs_difference_with_indexes(self, numbers):
        if len(numbers) < 2:
            raise ValueError("The list must contain at least two elements.")

        # Store the original indexes
        indexed_numbers = list(enumerate(numbers))

        # Sort the list based on values while keeping the original indexes
        sorted_numbers = sorted(indexed_numbers, key=lambda x: x[1])

        # Initialize variables to track the minimum difference and indexes
        min_diff = float('inf')
        min_indexes = (-1, -1)

        # Calculate the absolute differences between consecutive elements
        for i in range(len(sorted_numbers) - 1):
            diff = abs(sorted_numbers[i][1] - sorted_numbers[i + 1][1])
            if diff < min_diff:
                min_diff = diff
                min_indexes = (sorted_numbers[i][0], sorted_numbers[i + 1][0])

        return min_diff, min_indexes

    # Function to plot the dataframe
    def plot_df(self, df, item, folder_name="plots", min_limit=0.4, max_limit=0.6):

        if not self.check_regression_in_range(df, min_limit, max_limit):
            return None, None, None, None

        x = df['index']

        # Calculate regression lines
        b_linear = np.poly1d(np.polyfit(x, df['b'], 1))(x)
        b_poly = np.poly1d(np.polyfit(x, df['b'], 2))(x)
        s_linear = np.poly1d(np.polyfit(x, df['s'], 1))(x)
        s_poly = np.poly1d(np.polyfit(x, df['s'], 2))(x)

        if b_poly[1] > b_poly[0]:
            optimum_b = max(b_poly)
        else:
            optimum_b = min(b_poly)

        if s_poly[1] > s_poly[0]:
            optimum_s = max(s_poly)
        else:
            optimum_s = min(s_poly)

        # small_limit = round(min(optimum_b, optimum_s), 4)
        # large_limit = round(max(optimum_b, optimum_s), 4)
        final_optimum = round(abs(optimum_b - optimum_s), 4)

        # Analyze relationships between lines
        b_max_distance = max(abs(b_linear - b_poly))
        s_max_distance = max(abs(s_linear - s_poly))

        label = "Default"
        if b_max_distance <= 0.03 and s_max_distance <= 0.03:
            # Check if b lines intersect s lines
            b_min = min(min(b_linear), min(b_poly))
            b_max = max(max(b_linear), max(b_poly))
            s_min = min(min(s_linear), min(s_poly))
            s_max = max(max(s_linear), max(s_poly))

            if b_max < s_min or s_max < b_min:
                label = "Parallel"
            else:
                intersections = 0
                for i in range(1, len(x)):
                    if b_linear[i - 1] <= s_linear[i - 1] and b_linear[i] > s_linear[i]:
                        intersections += 1
                    if b_linear[i - 1] >= s_linear[i - 1] and b_linear[i] < s_linear[i]:
                        intersections += 1
                    if b_poly[i - 1] <= s_poly[i - 1] and b_poly[i] > s_poly[i]:
                        intersections += 1
                    if b_poly[i - 1] >= s_poly[i - 1] and b_poly[i] < s_poly[i]:
                        intersections += 1
                    if b_linear[i - 1] <= s_poly[i - 1] and b_linear[i] > s_poly[i]:
                        intersections += 1
                    if s_linear[i - 1] <= b_poly[i - 1] and s_linear[i] > b_poly[i]:
                        intersections += 1

                if intersections <= 3:
                    label = "One Point"
                elif intersections > 3:
                    label = "Two Points"
        else:
            # Check if b lines intersect s lines
            b_min = min(min(b_linear), min(b_poly))
            b_max = max(max(b_linear), max(b_poly))
            s_min = min(min(s_linear), min(s_poly))
            s_max = max(max(s_linear), max(s_poly))

            if b_max < s_min or s_max < b_min:
                label = "Default"
            else:
                intersections = 0
                for i in range(1, len(x)):
                    if b_linear[i - 1] <= s_linear[i - 1] and b_linear[i] > s_linear[i]:
                        intersections += 1
                    if b_linear[i - 1] >= s_linear[i - 1] and b_linear[i] < s_linear[i]:
                        intersections += 1
                    if b_poly[i - 1] <= s_poly[i - 1] and b_poly[i] > s_poly[i]:
                        intersections += 1
                    if b_poly[i - 1] >= s_poly[i - 1] and b_poly[i] < s_poly[i]:
                        intersections += 1
                    if b_linear[i - 1] <= s_poly[i - 1] and b_linear[i] > s_poly[i]:
                        intersections += 1
                    if s_linear[i - 1] <= b_poly[i - 1] and s_linear[i] > b_poly[i]:
                        intersections += 1

                if intersections > 3:
                    label = "Two Points"

        if label:
            main_path = os.path.join(self.folder_path, folder_name) if folder_name else self.folder_path
            if not os.path.exists(main_path):
                os.makedirs(main_path)

            child_folder_1 = f"[{min_limit}_{max_limit}]"
            child_folder_2 = f"[{str(df.iloc[0]['index'])}_{str(df.iloc[len(df) - 1]['index'])}]_{str(df.iloc[len(df) - 1]['Time']).replace(':', ';')[0:-3] + '_' + str(df.iloc[len(df) - 1]['Date'])}"
            save_path = os.path.join(main_path, child_folder_1, label, child_folder_2)

            fig_name = f"{save_path}/[{item[0]}_{item[-1]}]_{df.iloc[-1]['Date']}_{str(df.iloc[-1]['Time']).replace(':', ';')[0:-3]}_[{min_limit}_{max_limit}].png"

            x_fit = np.linspace(0, len(b_linear) - 1, len(b_linear))
            all_y_fits = np.vstack((b_linear, b_poly, s_linear, s_poly))

            # Calculate max and min across all y_fit arrays
            max_y_fit = np.max(all_y_fits, axis=0)
            min_y_fit = np.min(all_y_fits, axis=0)
            area = np.trapezoid(max_y_fit - min_y_fit, x_fit)

            return child_folder_2, final_optimum, fig_name, area
        return None, None, None, None

    def start(self):
        if not self.file_path1 or not self.folder_path:
            messagebox.showerror("Error", "Please select an Excel file and an output folder.")
            return

        # Set the logging configuration
        logging.basicConfig(filename="file.log",
                            filemode='a',
                            level=logging.DEBUG,
                            format='%(levelname)s | %(asctime)s | %(message)s')

        # Continue with existing start method logic
        if self.min_limit.get() != "":
            min_limit = self.min_limit.get().split("_")
            min_limit_list = [float(x) for x in min_limit]
        else:
            min_limit_list = [0, 0.1, 0.2, 0.3, 0.4]
        if self.max_limit.get() != "":
            max_limit = self.max_limit.get().split("_")
            max_limit_list = [float(x) for x in max_limit]
        else:
            max_limit_list = [1, 0.9, 0.8, 0.7, 0.6]
        if self.folder_name.get() != "":
            folder_name = self.folder_name.get()

        if self.folder_name.get() == "":
            folder_name = "plots"

        if self.data_rows.get() != "":
            data_rows = int(self.data_rows.get())

        try:
            df = pd.read_excel(self.file_path1, header=None)
            main_path = os.path.join(self.folder_path, folder_name) if folder_name else self.folder_path
            if not os.path.exists(main_path):
                os.makedirs(main_path)

            df = df.iloc[:, :6]

            df.columns = ['Date', 'Time', 'O', 'H', 'L', 'C']
            df['index'] = df.index
            df['index'] = range(1, len(df) + 1)
            df = df[['index', 'Date', 'Time', 'O', 'H', 'L', 'C']]
            df["Time"] = df["Time"].astype(str).str.strip()

            df['b'] = (df['C'] - df['L']) / (df['H'] - df['L'])
            df['s'] = (df['H'] - df['C']) / (df['H'] - df['L'])
        except Exception as e:
            logging.error(f"{inspect.currentframe().f_lineno}  | An error occurred while reading the Excel file: {e}")
            messagebox.showerror("Error", "An error occurred while reading the Excel file.")
            return

        r_e_dict = {}
        positive_negative = {}
        pivot_dict = {}
        after_pivot = {}
        logging.info(f"Start processing the abs for these {len(df) - int(self.start_row.get())} number of table.")
        table_counter = 0
        new_selected_time_indexes = list(range(int(self.start_row.get())-1, data_rows))
        date_name_value_list = []
        time_name_value_list = []

        if self.rows_number.get() != "":
            rows_number = int(self.rows_number.get()) + 1
        else:
            rows_number = 0

        for index in new_selected_time_indexes:
            try:
                sub_df_row = df.loc[rows_number:index]
                sub_df_reset = sub_df_row.reset_index(drop=True)
                sub_df = sub_df_reset.drop("index", axis=1)
                sub_df['index'] = range(1, len(sub_df) + 1)
                sub_df = sub_df[['index', 'Date', 'Time', 'O', 'H', 'L', 'C', 'b', 's']]
                last_date_sub_df = sub_df.iloc[-1]['Date'].replace(".", "-")
                last_time_sub_df = sub_df.iloc[-1]['Time']
                date_name_value_list.append(last_date_sub_df)
                time_name_value_list.append(last_time_sub_df)
            except Exception as e:
                logging.error(f"{inspect.currentframe().f_lineno}  | An error occurred while processing the data: {e}")
                messagebox.showerror("Error", "An error occurred while processing the data.")
                return

            df_2_dict = {}
            for i in range(len(min_limit_list)):
                all_dates = []
                all_times = []
                all_density_b = []
                all_density_s = []
                all_measure = []
                all_color = []
                temp_areas = []
                all_points = []

                first_min_limit = min_limit_list[i]
                first_max_limit = max_limit_list[i]
                df_name = f"df_{str(first_min_limit)}_{str(first_max_limit)}"

                start = len(sub_df) + 1
                end = len(sub_df)
                try:
                    sequences = self.print_sequences(start, end)
                except Exception as e:
                    logging.error(
                        f"{inspect.currentframe().f_lineno}  | An error occurred while generating sequences: {e}")
                    messagebox.showerror("Error", "An error occurred while generating sequences.")
                    return

                for item in sequences:
                    temp_df = sub_df.loc[sub_df['index'].isin(item)]

                    if len(temp_df) > len(sub_df):
                        continue
                    try:
                        child_folder, final_difference, fig_name, area = self.plot_df(temp_df, item, folder_name,
                                                                                      min_limit=first_min_limit,
                                                                                      max_limit=first_max_limit)
                    except Exception as e:
                        logging.error(
                            f"{inspect.currentframe().f_lineno}  | An error occurred while self function: {e}")
                        messagebox.showerror("Error", "An error occurred while self function.")
                        return
                    if child_folder and final_difference:
                        temp_date = temp_df.iloc[-1]['Date']
                        temp_time = temp_df.iloc[-1]['Time']
                        temp_density_b = round(temp_df.iloc[-1]['b'], 6)
                        temp_density_s = round(temp_df.iloc[-1]['s'], 6)
                        temp_measure = round(temp_df.iloc[-1]['H'] - temp_df.iloc[-1]['L'], 3)
                        temp_color = round(temp_df.iloc[-1]['C'] - temp_df.iloc[-1]['O'], 3)
                        all_dates.append(temp_date)
                        all_times.append(temp_time)
                        all_density_b.append(temp_density_b)
                        all_density_s.append(temp_density_s)
                        all_measure.append(temp_measure)
                        all_color.append(temp_color)
                        temp_areas.append(area / len(temp_df))
                        all_points.append(len(temp_df))

                all_df = pd.DataFrame(
                    {'Date': all_dates, 'Time': all_times, 'Density_b': all_density_b, 'Density_s': all_density_s,
                     'Measure': all_measure, 'Color': all_color, 'Special': temp_areas, 'Points': all_points})

                column_name = str(first_min_limit) + '_' + str(first_max_limit)

                if len(all_df) == 0:
                    empty_dataframe = pd.DataFrame(columns=['Date', 'Time', 'Color', column_name, 'r.e', 'Minimum_Area'])
                    df_2_dict[df_name] = empty_dataframe
                    continue

                try:
                    # Sort all_df by Date
                    all_df['Date'] = pd.to_datetime(all_df['Date'])
                    all_df = all_df.sort_values(by='Date')

                    all_df = all_df.sort_values(by='Date', ascending=False)
                    all_df['Date_Time'] = all_df['Date'].astype(str) + '_' + all_df['Time'].astype(str)
                    time_counts = all_df['Date_Time'].value_counts().sort_index()
                    all_df['Number_of_repeat'] = all_df['Date_Time'].map(time_counts)
                    df_grouped = all_df.groupby(['Time', 'Date'])['Special'].agg(Minimum_Area='min',
                                                                                 Total_Area='sum').reset_index()
                    df_merged = pd.merge(all_df, df_grouped, on=['Time', 'Date'], how='left')

                    df_merged = df_merged.groupby(['Date', 'Time']).apply(MyApp.custom_operations).reset_index(
                        drop=True)

                    df_unique = df_merged.drop_duplicates(subset='Date_Time').sort_values(by='Number_of_repeat',
                                                                                          ascending=False).reset_index(
                        drop=True)
                    df_unique = df_unique[
                        ['Date', 'Number_of_repeat', 'Time', 'Density_b', 'Density_s', 'Measure', 'Color',
                         'Minimum_Area',
                         'Total_Area', 'Not', 'OK']]
                    # df_unique = df_unique[['Date', 'Number_of_repeat', 'Time', 'Minimum_Area', 'Total_Area']]
                    df_sorted = df_unique.sort_values(by=['Date', 'Number_of_repeat'], ascending=[False, False])
                    df_sorted['Date'] = df_sorted['Date'].astype(str)

                    df_2 = df_sorted[['Date', 'Time', 'Color', 'Not', 'OK', 'Minimum_Area']]

                    df_2[column_name] = abs(df_2['Not'] - df_2['OK'])
                    df_2['r.e'] = df_2.apply(
                        lambda row: row[column_name] * (-1) if row['Color'] < 0 else row[column_name],
                        axis=1)

                    df_2 = df_2[['Date', 'Time', 'Color', column_name, 'r.e', 'Minimum_Area']]
                    df_2_dict[df_name] = df_2
                except Exception as e:
                    logging.error(
                        f"{inspect.currentframe().f_lineno}  | An error occurred while processing the data: {e}")
                    messagebox.showerror("Error", "An error occurred while processing the data.")
                    return

            # current_year = str(datetime.now().year)


            abstract_df = pd.DataFrame()
            pivot_date = []
            pivot_time = []
            pivot_total = []
            pivot_minimum_areas = []
            pivot_color = []
            for temp_excel in df_2_dict.values():
                minimum_row_temp = len(temp_excel)
                pivot_date.extend(temp_excel['Date'].tolist()[:minimum_row_temp])
                pivot_time.extend(temp_excel['Time'].tolist()[:minimum_row_temp])
                pivot_total.extend(temp_excel.iloc[:, 3].tolist()[:minimum_row_temp])
                pivot_minimum_areas.extend(temp_excel['Minimum_Area'].tolist()[:minimum_row_temp])
                pivot_color.extend(temp_excel['Color'].tolist()[:minimum_row_temp])
                try:
                    abstract_df = pd.concat([abstract_df, temp_excel.reset_index(drop=True)], axis=1)
                except Exception as e:
                    logging.error(
                        f"{inspect.currentframe().f_lineno}  | An error occurred while concatenating the data: {e}")
                    messagebox.showerror("Error", "An error occurred while concatenating the data.")
                    return

            try:
                pivot_df = pd.DataFrame({
                    'date': pivot_date,
                    'time': pivot_time,
                    'total': pivot_total,
                    'minimum_area': pivot_minimum_areas,
                    'color': pivot_color
                })

                pivot_df['date_time'] = pivot_df['date'] + "_" + pivot_df['time']
                grouped_pivot_df = pivot_df.groupby('date_time').agg(
                    {'total': 'sum', 'time': lambda x: ', '.join(set(x)), 'date_time': 'count',
                     'date': lambda x: ', '.join(set(x)), 'minimum_area': 'mean',
                     'color': 'mean'}).rename(columns={'date_time': 'num rep'})
                # sort by total
                grouped_pivot_df = grouped_pivot_df.sort_values(by='total', ascending=False)

                grouped_pivot_df = grouped_pivot_df.reset_index()

                grouped_pivot_df = grouped_pivot_df[['date', 'time', 'total', 'num rep', 'minimum_area', 'color']]
            except Exception as e:
                logging.error(
                    f"{inspect.currentframe().f_lineno}  | An error occurred while processing the pivot df: {e}")
                messagebox.showerror("Error", "An error occurred while processing the pivot df.")
                return

            abs_column_names = ["row_id", "row", "date", "time"]
            for i in range(len(min_limit_list)):
                abs_column_names.append(f"{min_limit_list[i]}--{max_limit_list[i]}")

            abs_column_names.append("total columns")
            r_e_list = []
            color_df = abstract_df[['Color']]
            date_df_check = abstract_df['Date']

            final_positive_value = False
            final_negative_value = False
            one_before_final_positive = False
            one_before_final_negative = False
            final_positive_value_list = []
            final_negative_value_list = []
            for i in range(len(color_df.columns)):
                if i == len(color_df.columns) - 1:
                    try:
                        temp_color_df_dropna = color_df.iloc[:, i].dropna()
                        if len(temp_color_df_dropna) > 0:
                            final_positive_value = sum(temp_color_df_dropna >= 0) / len(temp_color_df_dropna)
                            final_negative_value = 1 - final_positive_value
                        if (final_positive_value != False) and (final_negative_value != False):
                            final_positive_value_list.append(final_positive_value)
                            final_negative_value_list.append(final_negative_value)
                    except Exception as e:
                        logging.error(
                            f"{inspect.currentframe().f_lineno}  | An error occurred while checking the final pivot data: {e}")
                        messagebox.showerror("Error", "An error occurred while checking the final pivot data.")
                        return
                else:
                    try:
                        temp_color_df_dropna = color_df.iloc[:, i].dropna()
                        if len(temp_color_df_dropna) > 0:
                            one_before_final_positive = sum(temp_color_df_dropna >= 0) / len(temp_color_df_dropna)
                            one_before_final_negative = 1 - one_before_final_positive

                        if (one_before_final_positive != False) and (one_before_final_negative != False):
                            final_positive_value_list.append(one_before_final_positive)
                            final_negative_value_list.append(one_before_final_negative)
                    except Exception as e:
                        logging.error(
                            f"{inspect.currentframe().f_lineno}  | An error occurred while checking the pivot data: {e}")
                        messagebox.showerror("Error", "An error occurred while checking the pivot data.")
                        return

            re_df = abstract_df[['r.e']]

            for i in range(len(re_df.columns)):
                if len(re_df.iloc[:, i].dropna()) == 0:
                    r_e_list.append("No Shape")
                else:
                    r_e_list.append(re_df.iloc[:, i].sum())

            r_e_list.append(sum(item for item in r_e_list if isinstance(item, (int, float, np.integer))))
            key_name = str(index+1)
            r_e_dict[key_name] = r_e_list
            try:
                if (final_positive_value != False) and (one_before_final_positive != False):
                    if len(final_positive_value_list) > 1 and len(final_negative_value_list) > 1:
                        positive = final_positive_value_list[-1] - final_positive_value_list[-2]
                        negative = final_negative_value_list[-1] - final_negative_value_list[-2]
                        positive_negative[key_name] = [round(positive, 3),
                                                       round(negative, 3)]
                    elif len(final_positive_value_list) == 1 and len(final_negative_value_list) == 1:
                        positive = final_positive_value_list[-1]
                        negative = final_negative_value_list[-1]
                        positive_negative[key_name] = [round(positive, 3),
                                                       round(negative, 3)]
                    else:
                        positive_negative[key_name] = [0, 0]
                elif final_positive_value != False:
                    positive = final_positive_value_list[-1]
                    negative = final_negative_value_list[-1]
                    positive_negative[key_name] = [round(positive, 3),
                                                   round(negative, 3)]
                else:
                    positive_negative[key_name] = [0, 0]
            except Exception as e:
                logging.error(
                    f"{inspect.currentframe().f_lineno}  | An error occurred while checking the positive and negative values: {e}")
                messagebox.showerror("Error", "An error occurred while checking the positive and negative values.")
                return

            pivot_list = []
            max_of_max_pivots = set()
            try:
                for i in range(len(grouped_pivot_df) - 1):
                    if grouped_pivot_df['total'].iloc[i] == grouped_pivot_df['total'].iloc[i + 1]:
                        pivot_list.append(grouped_pivot_df['total'].iloc[i])
                    max_of_max_pivots.add(max(grouped_pivot_df['total']))
            except Exception as e:
                logging.error(
                    f"{inspect.currentframe().f_lineno}  | An error occurred while checking the pivot list: {e}")
                messagebox.showerror("Error", "An error occurred while checking the pivot list.")
                return

            try:
                counts = Counter(pivot_list)
                for number, count in counts.items():
                    if count > 1:
                        pivot_list.append(number)
                pivot_list.sort(reverse=True)
                max_pivot = max(max_of_max_pivots)
            except Exception as e:
                logging.error(
                    f"{inspect.currentframe().f_lineno}  | An error occurred while checking the pivot list: {e}")
                messagebox.showerror("Error", "An error occurred while checking the pivot list.")
                return

            counts = Counter(pivot_list)

            result = []
            check_temp_minimum_area = 1000
            try:
                if len(counts.items()) > 0:
                    for number, count in counts.items():
                        selected_pivot_df = grouped_pivot_df[grouped_pivot_df['total'] == number]
                        color_list = grouped_pivot_df[grouped_pivot_df['total'] == number]['color'].tolist()
                        cell_color = self.check_numbers(color_list)

                        minimum_area_list = grouped_pivot_df[grouped_pivot_df['total'] == number][
                            'minimum_area'].tolist()
                        temp_minimum_area, temp_min_area_indexes = self.find_min_abs_difference_with_indexes(
                            minimum_area_list)
                        temp_after_pivot = []
                        if temp_minimum_area < check_temp_minimum_area:
                            check_temp_minimum_area = temp_minimum_area
                            time_first_minimum_area = selected_pivot_df['time'].iloc[temp_min_area_indexes[1]]
                            time_second_minimum_area = selected_pivot_df['time'].iloc[temp_min_area_indexes[0]]
                            date_first_minimum_area = selected_pivot_df['date'].iloc[temp_min_area_indexes[1]]
                            date_second_minimum_area = selected_pivot_df['date'].iloc[temp_min_area_indexes[0]]
                            minidef_pivot_color_first = selected_pivot_df['color'].iloc[temp_min_area_indexes[0]]
                            minidef_pivot_color_second = selected_pivot_df['color'].iloc[temp_min_area_indexes[1]]
                            both_minidef_pivot_color = self.check_numbers(
                                [minidef_pivot_color_first, minidef_pivot_color_second])

                        if count > 1:
                            result.append(f"{cell_color}_{number}_" + "_".join([str(number)] * (count - 1)))
                        else:
                            result.append(f"{cell_color}_{str(number)}")

                        temp_after_pivot.append(
                            f"{both_minidef_pivot_color}_{time_first_minimum_area}_{time_second_minimum_area}")
                        temp_after_pivot.append(f'{both_minidef_pivot_color}_{date_first_minimum_area}_{date_second_minimum_area}')
                        temp_after_pivot.append(f"{both_minidef_pivot_color}_{round(check_temp_minimum_area, 4)}")
                        temp_after_pivot.append(max_pivot)
                        temp_after_pivot.append(sum(grouped_pivot_df['total']))
                else:
                    temp_after_pivot = ["No Time", "No Minimum Area", "No Minimum Area", max_pivot,
                                        sum(grouped_pivot_df['total'])]
            except Exception as e:
                logging.error(
                    f"{inspect.currentframe().f_lineno}  | An error occurred while checking the pivot list counts: {e}")
                messagebox.showerror("Error", "An error occurred while checking the pivot list counts.")
                return

            after_pivot[key_name] = temp_after_pivot
            pivot_dict[key_name] = result
            table_counter += 1
            logging.info(
                f"{inspect.currentframe().f_lineno}  | The table {table_counter} is from 0 - {index+1} done successfully.")
            logging.info("----------------------------------------------------------------------------------------------------------------------------------------------------")

        max_len_piv_list = []
        for item in pivot_dict.values():
            max_len_piv_list.append(len(item))

        maximum_value = max(max_len_piv_list)

        try:
            for i in range(maximum_value):
                abs_column_names.append(f"pivot_{i + 1}")

            abs_column_names.append("minidef pivote")
            abs_column_names.append("minimu def area date")
            abs_column_names.append("minimu def area")
            abs_column_names.append("max pivot")
            abs_column_names.append("total pivot")
            abs_column_names.append("positive")
            abs_column_names.append("negative")
        except Exception as e:
            logging.error(
                f"{inspect.currentframe().f_lineno}  | An error occurred while creating the column names: {e}")
            messagebox.showerror("Error", "An error occurred while creating the column names.")
            return

        try:
            final_pivot_dict = {}
            for i, j in pivot_dict.items():
                pippo = j
                for _ in range(maximum_value - len(j)):
                    pippo.extend(".")
                final_pivot_dict[i] = pippo
        except Exception as e:
            logging.error(
                f"{inspect.currentframe().f_lineno}  | An error occurred while creating the final pivot dict: {e}")
            messagebox.showerror("Error", "An error occurred while creating the final pivot dict.")
            return

        final_list = {}
        for i in range(len(r_e_dict)):
            key_name = (list(positive_negative.keys())[i])
            pippo_list = [key_name]
            pippo_list.extend([date_name_value_list[i]])
            pippo_list.extend([str(time_name_value_list[i])])
            pippo_list.extend((list(r_e_dict.values())[i]))
            pippo_list.extend((list(final_pivot_dict.values())[i]))
            pippo_list.extend((list(after_pivot.values())[i]))
            pippo_list.extend((list(positive_negative.values())[i]))

            final_list[f"row_{i + 1}"] = pippo_list

        final_abs_df = pd.DataFrame.from_dict(final_list, orient='index')

        # Optionally, reset the index and assign proper column names
        final_abs_df.reset_index(inplace=True)

        final_abs_df.columns = abs_column_names

        sorted_final_abs_df = final_abs_df[abs_column_names[1:]]
        # sorted_final_abs_df['date'] = pd.to_datetime(final_abs_df['date'])

        # # Sort the DataFrame by 'date' column from newest to oldest
        # sorted_final_abs_df = final_abs_df.sort_values(by='date', ascending=False)
        # sorted_final_abs_df['date'] = sorted_final_abs_df['date'].dt.strftime('%Y-%m-%d')

        # try:
        #     second_df = pd.read_excel(self.file_path1, header=None)
        #     if second_df.empty:
        #         logging.error("There is no data in the excel file.")
        #     else:
        #         pattern = r'\b\d{4}\.\d{2}\.\d{2}\b'
        #         matches = re.findall(pattern, second_df[0].iloc[0])
        #         if len(matches) == 0:
        #             second_df = second_df.drop(0).reset_index(drop=True)
        #         new_header = ['date', 'time', 'open', 'high', 'low', 'close']
        #         second_df.columns = new_header
        #         second_df['date'] = second_df['date'].astype(str)
        #         second_df['time'] = second_df['time'].astype(str)
        #
        # except Exception as e:
        #     logging.error(f":There is an error in reading excel file: {e}")

        # if second_df is not None:
        #     results = self.analyze_data(second_df)
        #     if results:
        #         column_names = ['date',
        #                         'ratio',
        #                         'PDP',
        #                         'chord slope',
        #                         'mid result',
        #                         'f result',
        #                         'C 14:30',
        #                         'C-O 15:30',
        #                         'C-O 16:30',
        #                         'C-O 17:30',
        #                         'C-O 18:30',
        #                         'C-O 19:30',
        #                         'C-O 20:30',
        #                         'C-O 21:30',
        #                         'C-O 22:30',
        #                         'C-O 23:30',
        #                         'point1',
        #                         'point2',
        #                         'point3',
        #                         'area1',
        #                         'area2',
        #                         'min',
        #                         'max',
        #                         'YPC']
        #         result_df = pd.DataFrame(results, columns=column_names)
        #         result_df.fillna(0, inplace=True)
        #
        #
        # result_df['date'] = result_df['date'].astype(str).str.replace(".", "-")
        # result_df['ratio'] = result_df.apply(self.update_ratio, axis=1)

        # merged_df = pd.merge(sorted_final_abs_df, result_df, on="date", how="outer")
        fillna_value = "."
        # merged_df_cleaned = merged_df.fillna(f"{fillna_value}")
        sorted_merged_df_cleaned = sorted_final_abs_df.fillna(f"{fillna_value}")

        # merged_df_cleaned = merged_df.dropna(subset=['positive', 'negative'], how='all')

        # merged_df_cleaned['date'] = pd.to_datetime(merged_df_cleaned['date'])

        # Sort the DataFrame by 'date' column from newest to oldest
        # sorted_merged_df_cleaned = merged_df_cleaned.sort_values(by='date', ascending=False)
        # sorted_merged_df_cleaned['date'] = sorted_merged_df_cleaned['date'].dt.strftime('%Y-%m-%d')

        sorted_merged_df_cleaned.to_excel(f"{main_path}\\final_abs.xlsx", index=False)

        # abstract_file_name = f"{sorted_path}\\abstract_{df_counter}.xlsx"
        # pivot_file_name = f"{sorted_path}\\pivot_{df_counter}.xlsx"
        # abstract_df.to_excel(abstract_file_name, sheet_name='abstract', index=False)
        # grouped_pivot_df.to_excel(pivot_file_name, sheet_name='pivot', index=False)

        file_path = f"{main_path}\\final_abs.xlsx"
        from datetime import datetime
        now = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")

        if len(sorted_final_abs_df) > 0:
            save_file_path = f"{main_path}\\abs_{int(self.start_row.get())}_{data_rows}_{now}.xlsx"
        else:
            save_file_path = f"{main_path}\\empty_abs.xlsx"

        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Define color fills
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
        purple_fill = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")  # Purple

        new_dataframe_columns = ['chord slope',
                            'mid result',
                            'f result',
                            'C-O 15:30',
                            'C-O 16:30',
                            'C-O 17:30',
                            'C-O 18:30',
                            'C-O 19:30',
                            'C-O 20:30',
                            'C-O 21:30',
                            'C-O 22:30',
                            'C-O 23:30']

        # Process each column
        for col in sheet.iter_cols(min_row=1):  # Start from the second row
            column_name = col[0].value
            if column_name == "date":
                continue  # Skip the 'date' column

            # Check column name patterns
            if "--" in column_name:
                for cell in col[1:]:
                    if cell.value != 'No Shape' and cell.value != fillna_value:
                        if cell.value < 0:
                            cell.fill = red_fill
                        elif cell.value > 0:
                            cell.fill = green_fill

            elif column_name.lower() == "total columns":
                for cell in col[1:]:
                    if cell.value != 'No Shape' and cell.value != fillna_value:
                        if cell.value < 0:
                            cell.fill = red_fill
                        elif cell.value > 0:
                            cell.fill = green_fill

            elif column_name == 'ratio':
                for cell in col[1:]:
                    if isinstance(cell.value, str) and "_" in cell.value and cell.value != fillna_value:
                        color, text = cell.value.split("_", 1)
                        if color == "Green":
                            cell.fill = green_fill
                        elif color == "Red":
                            cell.fill = red_fill
                        cell.value = text

            elif column_name.startswith(("pivot", "minidef", "minimu")):
                for cell in col[1:]:
                    if isinstance(cell.value, str) and "_" in cell.value and cell.value != fillna_value:
                        color, text = cell.value.split("_", 1)
                        if color == "Green":
                            cell.fill = green_fill
                        elif color == "Purple":
                            cell.fill = purple_fill
                        elif color == "Red":
                            cell.fill = red_fill
                        cell.value = text  # Update cell value to remove the prefix

            elif column_name.lower() in ("negative", "positive"):
                for cell in col[1:]:
                    if cell.value != fillna_value:
                        if cell.value < 0:
                            cell.fill = red_fill
                        elif cell.value > 0:
                            cell.fill = green_fill

            elif column_name in new_dataframe_columns:
                for cell in col[1:]:
                    if cell.value != 0 and cell.value != fillna_value:
                        if cell.value < 0:
                            cell.fill = red_fill
                        elif cell.value > 0:
                            cell.fill = green_fill

            elif column_name == 'YPC':
                for cell in col[1:]:
                    if cell.value != 0 and cell.value != fillna_value:
                        if cell.value < 0:
                            cell.fill = red_fill
                        elif cell.value > 0:
                            cell.fill = green_fill

            elif column_name == 'min':
                for cell in col[1:]:
                    if cell.value != fillna_value:
                        cell.fill = red_fill

            elif column_name == 'max':
                for cell in col[1:]:
                    if cell.value != fillna_value:
                        cell.fill = green_fill

        # Save the updated workbook
        workbook.save(save_file_path)
        os.remove(file_path)

        messagebox.showinfo("Process Completed", "The process has finished successfully!")


if __name__ == "__main__":
    # ask for password
    # if not MyApp().ask_password():
    #     exit()
    app = MyApp()
    app.mainloop()
